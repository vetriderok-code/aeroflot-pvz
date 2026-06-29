from django.test import SimpleTestCase
from io import BytesIO

from openpyxl import Workbook

from flights.utils.excel_settlement_report import (
    ReportColumnConfig,
    _find_table_layout,
    column_letter_to_index,
    index_to_column_letter,
    process_report_full,
    resolve_north_east_meters,
    sk42_meters_to_wgs84,
)


class ExcelSettlementReportTests(SimpleTestCase):
    def test_column_letters(self):
        self.assertEqual(column_letter_to_index('G'), 7)
        self.assertEqual(column_letter_to_index('H'), 8)
        self.assertEqual(column_letter_to_index('I'), 9)
        self.assertEqual(index_to_column_letter(7), 'G')

    def test_report_column_config_defaults(self):
        cfg = ReportColumnConfig.from_letters()
        self.assertEqual(cfg.east_col, 7)
        self.assertEqual(cfg.north_col, 8)
        self.assertEqual(cfg.settlement_col, 9)

    def test_resolve_swapped_g_h_values(self):
        # G=5359504 (фактически Y), H=7360000 (фактически X) — типичная сводная
        north, east = resolve_north_east_meters(5359504, 7360000)
        self.assertEqual(north, 5359504)
        self.assertEqual(east, 7360000)

    def test_resolved_coords_not_julitornio(self):
        north, east = resolve_north_east_meters(5359504, 7360000)
        lat, lon = sk42_meters_to_wgs84(north, east)
        self.assertIsNotNone(lat)
        self.assertLess(lat, 55)
        self.assertGreater(lon, 34)

    def test_naive_column_order_gives_wrong_latitude(self):
        lat_bad, _ = sk42_meters_to_wgs84(7360000, 5359504)
        self.assertGreater(lat_bad, 55)

    def test_explicit_columns_without_np_header(self):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        ws['C1'] = 'X'
        ws['D1'] = 'Y'
        ws['E1'] = 'НП'
        ws['C2'] = 7360000
        ws['D2'] = 5359504
        cfg = ReportColumnConfig.from_letters(north='D', east='C', settlement='E')
        layout = _find_table_layout(ws, cfg)
        self.assertIsNotNone(layout)
        found_cfg, data_start = layout
        self.assertEqual(found_cfg.east_col, 3)
        self.assertEqual(found_cfg.north_col, 4)
        self.assertEqual(found_cfg.settlement_col, 5)
        self.assertEqual(data_start, 2)

    def test_process_report_with_custom_columns(self):
        wb = Workbook()
        ws = wb.active
        ws['C2'] = 7360000
        ws['D2'] = 5359504
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        class Upload:
            name = 'test.xlsx'
            size = buf.getbuffer().nbytes

            def read(self):
                return buf.getvalue()

        cfg = ReportColumnConfig.from_letters(north='D', east='C', settlement='E')
        excel_bytes, kml_bytes, meta = process_report_full(Upload(), columns=cfg)
        self.assertGreater(meta.point_count, 0)
        self.assertTrue(excel_bytes)
        self.assertTrue(kml_bytes)
