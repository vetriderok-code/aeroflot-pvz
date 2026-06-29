import openpyxl
from django.test import SimpleTestCase

from flights.utils.excel_import_share import (
    find_svodnaya_sheet_name,
    load_calculation_pilot_mapping,
)


class FindSvodnayaSheetTests(SimpleTestCase):
    def test_prefers_exact_svodnaya_over_old(self):
        wb = openpyxl.Workbook()
        wb.active.title = 'сводная_old'
        wb.create_sheet('СВОДНАЯ')
        self.assertEqual(find_svodnaya_sheet_name(wb), 'СВОДНАЯ')

    def test_accepts_svodnaya_latin(self):
        wb = openpyxl.Workbook()
        wb.active.title = 'SVODNAYA'
        self.assertEqual(find_svodnaya_sheet_name(wb), 'SVODNAYA')

    def test_rejects_only_old_sheet(self):
        wb = openpyxl.Workbook()
        wb.active.title = 'сводная_old'
        self.assertIsNone(find_svodnaya_sheet_name(wb))


class LoadCalculationPilotMappingTests(SimpleTestCase):
    def test_reads_info_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Информация'
        ws.cell(3, 12, 'Каган')
        ws.cell(3, 13, 5)
        ws.cell(4, 12, 'Дюша')
        ws.cell(4, 13, 2)

        mapping = load_calculation_pilot_mapping(wb)
        self.assertEqual(mapping[5], 'Каган')
        self.assertEqual(mapping[2], 'Дюша')

    def test_missing_sheet_returns_empty(self):
        wb = openpyxl.Workbook()
        self.assertEqual(load_calculation_pilot_mapping(wb), {})
