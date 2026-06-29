"""Источник Excel для импорта: загрузка из HTTP или с диска (фоновый импорт)."""
from __future__ import annotations

import os
from pathlib import Path

import openpyxl


class UploadedExcelSource:
    def __init__(self, uploaded_file):
        self._file = uploaded_file

    @property
    def name(self) -> str:
        return self._file.name

    @property
    def size(self) -> int:
        return self._file.size

    def seek(self, pos: int) -> None:
        self._file.seek(pos)

    def read(self) -> bytes:
        return self._file.read()

    def open_workbook(self):
        self.seek(0)
        try:
            if self.name.endswith('.xlsm'):
                return openpyxl.load_workbook(self._file, data_only=True, keep_vba=True)
            return openpyxl.load_workbook(self._file, data_only=True)
        except Exception:
            self.seek(0)
            return openpyxl.load_workbook(self._file, data_only=True)


class PathExcelSource:
    def __init__(self, path: str | Path):
        self.path = str(Path(path).resolve())
        self.name = os.path.basename(self.path)
        self.size = os.path.getsize(self.path)

    def seek(self, pos: int) -> None:
        pass

    def read(self) -> bytes:
        with open(self.path, 'rb') as handle:
            return handle.read()

    def open_workbook(self):
        try:
            if self.name.endswith('.xlsm'):
                return openpyxl.load_workbook(self.path, data_only=True, keep_vba=True)
            return openpyxl.load_workbook(self.path, data_only=True)
        except Exception:
            return openpyxl.load_workbook(self.path, data_only=True)
