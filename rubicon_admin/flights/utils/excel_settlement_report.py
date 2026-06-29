"""Excel-отчёт: заполнение НП и выгрузка KML по координатам СК-42."""
from __future__ import annotations

import io
import logging
import re
from dataclasses import dataclass
from xml.sax.saxutils import escape

from openpyxl import load_workbook

from flights.models import Flight
from flights.utils.nearest_settlement import get_nearest_settlement_name

logger = logging.getLogger(__name__)

MAX_FILE_SIZE = 25 * 1024 * 1024


@dataclass
class SheetStats:
    name: str
    filled: int = 0
    skipped: int = 0
    errors: int = 0


@dataclass
class ReportPoint:
    sheet: str
    row: int
    seq: str = ''
    time: str = ''
    target: str = ''
    detail: str = ''
    section: str = ''
    x_sk42: float = 0.0
    y_sk42: float = 0.0
    lat: float | None = None
    lon: float | None = None
    settlement: str = ''


@dataclass
class FullProcessResult:
    excel_filename: str
    kml_filename: str
    sheets: list[SheetStats]
    total_filled: int = 0
    point_count: int = 0
    source_name: str = ''


@dataclass
class ProcessResult:
    output: io.BytesIO
    filename: str
    sheets: list[SheetStats]
    total_filled: int = 0


@dataclass
class KmlResult:
    output: io.BytesIO
    filename: str
    point_count: int = 0


@dataclass
class ReportColumnConfig:
    """Столбцы Excel: Y (север), X (восток), ближайший НП."""

    north_col: int
    east_col: int
    settlement_col: int

    @classmethod
    def from_letters(
        cls,
        north: str = 'H',
        east: str = 'G',
        settlement: str = 'I',
    ) -> 'ReportColumnConfig':
        return cls(
            north_col=column_letter_to_index(north),
            east_col=column_letter_to_index(east),
            settlement_col=column_letter_to_index(settlement),
        )


def column_letter_to_index(letter: str) -> int:
    text = re.sub(r'[^A-Za-z]', '', (letter or '').strip().upper())
    if not text:
        raise ValueError(f'Некорректный столбец: {letter!r}')
    index = 0
    for char in text:
        index = index * 26 + (ord(char) - ord('A') + 1)
    return index


def index_to_column_letter(index: int) -> str:
    if index < 1:
        raise ValueError(f'Некорректный индекс столбца: {index}')
    letters = []
    while index:
        index, rem = divmod(index - 1, 26)
        letters.append(chr(rem + ord('A')))
    return ''.join(reversed(letters))


def _parse_sk42_cell(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(',', '.')
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _cell_text(value) -> str:
    if value is None:
        return ''
    if hasattr(value, 'strftime'):
        return value.strftime('%H:%M')
    return str(value).strip()


def sk42_meters_to_wgs84(
    north_meters: float,
    east_meters: float,
) -> tuple[float | None, float | None]:
    """Север (Y) и восток (X) в метрах СК-42 → WGS84 (как на карте портала)."""
    flight = Flight(coordinates=f'{north_meters} {east_meters}')
    lat_sk, lon_sk = flight.parse_coordinates_sk42()
    if lat_sk is None or lon_sk is None:
        return None, None
    return flight.sk42_to_wgs84(lat_sk, lon_sk)


def _sk42_value_role(value: float) -> str | None:
    """Эвристика: 52–59xxxxx ≈ север (Y), 60–79xxxxx ≈ восток (X)."""
    if 5_200_000 <= value <= 5_999_999:
        return 'north'
    if 6_000_000 <= value <= 7_999_999:
        return 'east'
    if 4_000_000 <= value < 5_200_000:
        return 'north'
    return None


def resolve_north_east_meters(first: float, second: float) -> tuple[float, float]:
    """
    Определить север/восток по значениям в двух ячейках.
    В сводной подписи G/H часто перепутаны с фактическими Y/X.
    """
    role_first = _sk42_value_role(first)
    role_second = _sk42_value_role(second)

    if role_first == 'north' and role_second == 'east':
        return first, second
    if role_first == 'east' and role_second == 'north':
        return second, first
    if role_first == 'north':
        return first, second
    if role_second == 'north':
        return second, first
    if role_first == 'east':
        return second, first
    if role_second == 'east':
        return first, second
    return first, second


def _find_xy_columns_in_row(ws, row: int) -> tuple[int | None, int | None]:
    east_col = north_col = None
    max_col = min(ws.max_column or 1, 30)
    for col in range(1, max_col + 1):
        label = str(ws.cell(row, col).value or '').strip().upper()
        if label in ('Х', 'X'):
            east_col = col
        elif label in ('У', 'Y', 'Ю'):
            north_col = col
    return north_col, east_col


def _find_first_coord_row(
    ws,
    east_col: int,
    north_col: int,
    *,
    start_row: int = 1,
    max_scan: int = 500,
) -> int | None:
    """Первая строка, где в указанных столбцах есть обе координаты."""
    max_row = min(ws.max_row or start_row, start_row + max_scan - 1)
    for row in range(start_row, max_row + 1):
        first_m = _parse_sk42_cell(ws.cell(row, east_col).value)
        second_m = _parse_sk42_cell(ws.cell(row, north_col).value)
        if first_m is not None and second_m is not None:
            return row
    return None


def _find_table_layout(
    ws,
    columns: ReportColumnConfig | None = None,
) -> tuple[ReportColumnConfig, int] | None:
    if columns is not None:
        data_start = _find_first_coord_row(
            ws,
            columns.east_col,
            columns.north_col,
        )
        if data_start is None:
            return None
        return columns, data_start

    max_scan = min(ws.max_row or 1, 120)
    settlement_col = columns.settlement_col if columns else None
    north_col = columns.north_col if columns else None
    east_col = columns.east_col if columns else None
    header_row = None

    for row in range(1, max_scan + 1):
        for col in range(1, min(ws.max_column or 1, 30) + 1):
            val = ws.cell(row, col).value
            if not val:
                continue
            text = str(val).lower()
            if 'ближайш' in text and ('н.п' in text or 'насел' in text):
                header_row = row
                if settlement_col is None:
                    settlement_col = col
                break
        if header_row:
            break

    if not header_row or not settlement_col:
        return None

    labeled_north, labeled_east = _find_xy_columns_in_row(ws, header_row)
    if labeled_north is None or labeled_east is None:
        sub_north, sub_east = _find_xy_columns_in_row(ws, header_row + 1)
        labeled_north = labeled_north or sub_north
        labeled_east = labeled_east or sub_east

    if columns is None:
        if labeled_east is not None:
            east_col = labeled_east
        elif east_col is None:
            east_col = settlement_col - 2
        if labeled_north is not None:
            north_col = labeled_north
        elif north_col is None:
            north_col = settlement_col - 1

    if east_col is None:
        east_col = settlement_col - 2
    if north_col is None:
        north_col = settlement_col - 1

    layout = ReportColumnConfig(
        north_col=north_col,
        east_col=east_col,
        settlement_col=settlement_col,
    )
    data_start = header_row + 1

    sub_north, sub_east = _find_xy_columns_in_row(ws, header_row + 1)
    if sub_east is not None and sub_north is not None:
        data_start = header_row + 2

    return layout, data_start


def _load_workbook_bytes(uploaded_file):
    if uploaded_file.size > MAX_FILE_SIZE:
        raise ValueError(f'Файл слишком большой (макс. {MAX_FILE_SIZE // (1024 * 1024)} МБ)')

    name = uploaded_file.name or 'report.xlsx'
    if not name.lower().endswith(('.xlsx', '.xlsm')):
        raise ValueError('Поддерживаются только файлы .xlsx и .xlsm')

    data = uploaded_file.read()
    keep_vba = name.lower().endswith('.xlsm')
    try:
        wb = load_workbook(io.BytesIO(data), keep_vba=keep_vba)
    except Exception as exc:
        raise ValueError(f'Не удалось открыть файл Excel: {exc}') from exc

    return wb, name, data, keep_vba


def _collect_points(
    ws,
    fill_settlement: bool = False,
    columns: ReportColumnConfig | None = None,
) -> tuple[list[ReportPoint], SheetStats]:
    stats = SheetStats(name=ws.title)
    layout_info = _find_table_layout(ws, columns)
    points: list[ReportPoint] = []
    if not layout_info:
        return points, stats

    layout, data_start = layout_info
    max_row = ws.max_row or data_start

    for row in range(data_start, max_row + 1):
        first_raw = ws.cell(row, layout.east_col).value
        second_raw = ws.cell(row, layout.north_col).value
        if first_raw is None and second_raw is None:
            continue

        first_m = _parse_sk42_cell(first_raw)
        second_m = _parse_sk42_cell(second_raw)
        if first_m is None or second_m is None:
            stats.skipped += 1
            continue

        north_m, east_m = resolve_north_east_meters(first_m, second_m)

        try:
            lat_wgs, lon_wgs = sk42_meters_to_wgs84(north_m, east_m)
            if lat_wgs is None or lon_wgs is None:
                stats.errors += 1
                if fill_settlement:
                    ws.cell(row, layout.settlement_col).value = '—'
                continue

            settlement = ''
            if fill_settlement:
                settlement = get_nearest_settlement_name(lat_wgs, lon_wgs) or '—'
                ws.cell(row, layout.settlement_col).value = settlement
                stats.filled += 1
            else:
                existing = ws.cell(row, layout.settlement_col).value
                settlement = _cell_text(existing) if existing else ''

            points.append(ReportPoint(
                sheet=ws.title,
                row=row,
                seq=_cell_text(ws.cell(row, 1).value),
                time=_cell_text(ws.cell(row, 3).value),
                target=_cell_text(ws.cell(row, 4).value),
                detail=_cell_text(ws.cell(row, 5).value),
                section=_cell_text(ws.cell(row, 6).value),
                x_sk42=east_m,
                y_sk42=north_m,
                lat=lat_wgs,
                lon=lon_wgs,
                settlement=settlement,
            ))
        except Exception as exc:
            logger.warning('Row %s sheet %s: %s', row, ws.title, exc)
            stats.errors += 1
            if fill_settlement:
                ws.cell(row, layout.settlement_col).value = '—'

    return points, stats


def process_report_full(
    uploaded_file,
    columns: ReportColumnConfig | None = None,
) -> tuple[bytes, bytes, FullProcessResult]:
    """Один проход: заполнить НП в Excel и собрать точки для KML."""
    wb, name, _, keep_vba = _load_workbook_bytes(uploaded_file)

    sheets_stats: list[SheetStats] = []
    all_points: list[ReportPoint] = []

    for ws in wb.worksheets:
        if ws.sheet_state == 'hidden':
            continue
        points, sheet_stats = _collect_points(
            ws,
            fill_settlement=True,
            columns=columns,
        )
        all_points.extend(points)
        if sheet_stats.filled or sheet_stats.errors:
            sheets_stats.append(sheet_stats)

    if not all_points:
        wb.close()
        cols = columns or ReportColumnConfig.from_letters()
        raise ValueError(
            'Не найдено строк с координатами. '
            f'Проверьте столбцы Y (север)={index_to_column_letter(cols.north_col)}, '
            f'X (восток)={index_to_column_letter(cols.east_col)}, '
            f'НП={index_to_column_letter(cols.settlement_col)}.'
        )

    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    wb.close()

    base = re.sub(r'\.(xlsx|xlsm)$', '', name, flags=re.I)
    excel_filename = f'{base}_np.xlsx' if not keep_vba else f'{base}_np.xlsm'
    kml_filename = f'{base}.kml'
    kml_text = _build_kml(all_points, name)

    meta = FullProcessResult(
        excel_filename=excel_filename,
        kml_filename=kml_filename,
        sheets=sheets_stats,
        total_filled=sum(s.filled for s in sheets_stats),
        point_count=len(all_points),
        source_name=name,
    )
    return excel_buf.getvalue(), kml_text.encode('utf-8'), meta


def process_excel_upload(
    uploaded_file,
    columns: ReportColumnConfig | None = None,
) -> ProcessResult:
    wb, name, _, keep_vba = _load_workbook_bytes(uploaded_file)

    sheets_stats: list[SheetStats] = []
    for ws in wb.worksheets:
        if ws.sheet_state == 'hidden':
            continue
        _, sheet_stats = _collect_points(
            ws,
            fill_settlement=True,
            columns=columns,
        )
        if sheet_stats.filled or sheet_stats.errors:
            sheets_stats.append(sheet_stats)

    if not any(s.filled for s in sheets_stats):
        wb.close()
        cols = columns or ReportColumnConfig.from_letters()
        raise ValueError(
            'Не найдено строк с координатами. '
            f'Проверьте столбцы Y={index_to_column_letter(cols.north_col)}, '
            f'X={index_to_column_letter(cols.east_col)}, '
            f'НП={index_to_column_letter(cols.settlement_col)}.'
        )

    output = io.BytesIO()
    wb.save(output)
    wb.close()
    output.seek(0)

    base = re.sub(r'\.(xlsx|xlsm)$', '', name, flags=re.I)
    out_name = f'{base}_np.xlsx' if not keep_vba else f'{base}_np.xlsm'

    return ProcessResult(
        output=output,
        filename=out_name,
        sheets=sheets_stats,
        total_filled=sum(s.filled for s in sheets_stats),
    )


def _build_kml(points: list[ReportPoint], source_name: str) -> str:
    title = escape(source_name)
    lines = [
        '<?xml version="1.0" encoding="UTF-8"?>',
        '<kml xmlns="http://www.opengis.net/kml/2.2">',
        '  <Document>',
        f'    <name>{title}</name>',
        f'    <description>Экспорт {len(points)} точек из Excel-отчёта</description>',
        '    <Style id="point">',
        '      <IconStyle>',
        '        <color>ff0000ff</color>',
        '        <scale>1.1</scale>',
        '        <Icon><href>http://maps.google.com/mapfiles/kml/shapes/placemark_circle.png</href></Icon>',
        '      </IconStyle>',
        '    </Style>',
    ]

    for idx, pt in enumerate(points, start=1):
        if pt.lat is None or pt.lon is None:
            continue
        if pt.lat == 90.0 and pt.lon == 0.0:
            continue

        label = pt.seq or str(idx)
        if pt.time:
            label = f'{label} {pt.time}'

        if not pt.settlement:
            pt.settlement = get_nearest_settlement_name(pt.lat, pt.lon) or '—'

        desc_rows = [
            ('Лист', pt.sheet),
            ('Строка', str(pt.row)),
            ('Время', pt.time),
            ('Характер цели', pt.target),
            ('Уточнение', pt.detail),
            ('Участок', pt.section),
            ('СК-42 Y (север)', str(int(pt.y_sk42))),
            ('СК-42 X (восток)', str(int(pt.x_sk42))),
            ('Ближайший НП', pt.settlement),
        ]
        desc_html = '<![CDATA[<table border="1" cellpadding="3" cellspacing="0">'
        for key, val in desc_rows:
            if val:
                desc_html += f'<tr><td><b>{escape(key)}</b></td><td>{escape(val)}</td></tr>'
        desc_html += '</table>]]>'

        lines.extend([
            '    <Placemark>',
            f'      <name>{escape(label.strip() or f"Точка {idx}")}</name>',
            f'      <description>{desc_html}</description>',
            '      <styleUrl>#point</styleUrl>',
            '      <Point>',
            f'        <coordinates>{pt.lon},{pt.lat},0</coordinates>',
            '      </Point>',
            '    </Placemark>',
        ])

    lines.extend(['  </Document>', '</kml>'])
    return '\n'.join(lines)


def generate_kml_from_excel(
    uploaded_file,
    columns: ReportColumnConfig | None = None,
) -> KmlResult:
    wb, name, _, _ = _load_workbook_bytes(uploaded_file)

    all_points: list[ReportPoint] = []
    for ws in wb.worksheets:
        if ws.sheet_state == 'hidden':
            continue
        points, _ = _collect_points(ws, fill_settlement=False, columns=columns)
        all_points.extend(points)

    wb.close()

    if not all_points:
        cols = columns or ReportColumnConfig.from_letters()
        raise ValueError(
            'Не найдено строк с координатами для KML. '
            f'Проверьте столбцы Y={index_to_column_letter(cols.north_col)}, '
            f'X={index_to_column_letter(cols.east_col)}.'
        )

    kml_text = _build_kml(all_points, name)
    output = io.BytesIO(kml_text.encode('utf-8'))
    output.seek(0)

    base = re.sub(r'\.(xlsx|xlsm)$', '', name, flags=re.I)
    return KmlResult(output=output, filename=f'{base}.kml', point_count=len(all_points))
