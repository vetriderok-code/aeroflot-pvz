"""Импорт пилотов и дежурств из листа «Расчёты» (Расчёты.xlsx)."""
from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime, time
from django.contrib.auth.models import Group
from django.db import transaction
from openpyxl import load_workbook

from flights.models import (
    Drone,
    OperatorCommLink,
    OperatorLocation,
    OperatorPlacementZone,
    OperatorProfile,
    Pilot,
    User,
)
from flights.utils.operator_dashboard import normalize_comm_links

SHEET_NAME = 'Расчёты'
SKIP_OPERATOR_NAMES = frozenset({'ОТРЫВ'})

DAY_SHIFT_START = time(6, 0)
DAY_SHIFT_END = time(18, 0)
NIGHT_SHIFT_START = time(18, 0)
NIGHT_SHIFT_END = time(6, 0)

_SHIFT_MAP = {
    'день': OperatorPlacementZone.DAY,
    'дневная': OperatorPlacementZone.DAY,
    'дневная смена': OperatorPlacementZone.DAY,
    'ночь': OperatorPlacementZone.NIGHT,
    'ночная': OperatorPlacementZone.NIGHT,
    'ночная смена': OperatorPlacementZone.NIGHT,
    'отрыв': OperatorPlacementZone.DETACHMENT,
}

_TIME_RANGE_RE = re.compile(
    r'(\d{1,2})[:\.](\d{2})\s*[-–—]\s*(\d{1,2})[:\.](\d{2})',
)


@dataclass(frozen=True)
class RaschetyRow:
    callname: str
    engineer_callname: str
    placement_zone: str | None
    duty_start: time | None
    duty_end: time | None
    comm_link_raw: str
    drone_raw: str
    settlement: str
    location_name: str
    group_name: str | None


@dataclass(frozen=True)
class RaschetyLocation:
    name: str
    settlement: str
    comm_links: tuple[str, ...]


@dataclass
class ImportStats:
    rows_read: int = 0
    pilots_created: int = 0
    pilots_updated: int = 0
    profiles_created: int = 0
    profiles_updated: int = 0
    locations_deleted: int = 0
    locations_created: int = 0
    locations_updated: int = 0
    profiles_deactivated: int = 0
    groups_assigned: int = 0
    groups_cleared: int = 0
    users_in_groups: int = 0
    pilot_locations_assigned: int = 0
    pilot_locations_cleared: int = 0
    drones_matched: int = 0
    drones_skipped: int = 0
    skipped: int = 0


_COMM_LINK_TOKEN_MAP = {
    'starlink': OperatorCommLink.STARLINK,
    'старлинк': OperatorCommLink.STARLINK,
    'star link': OperatorCommLink.STARLINK,
    'радиостанция старлинк': OperatorCommLink.STARLINK,
    'bshpd': OperatorCommLink.BSHPD,
    'бшпд': OperatorCommLink.BSHPD,
    'optics': OperatorCommLink.OPTICS,
    'оптика': OperatorCommLink.OPTICS,
    'lte': OperatorCommLink.LTE,
    'sks': OperatorCommLink.SKS,
    'скс': OperatorCommLink.SKS,
    'rj45': OperatorCommLink.RJ45,
    'p274m': OperatorCommLink.P274M,
    'п-274м': OperatorCommLink.P274M,
    'п274м': OperatorCommLink.P274M,
}


def parse_comm_links_from_cell(value) -> list[str]:
    text = _normalize_text(value)
    if not text:
        return []
    lowered = text.lower()
    if lowered in _COMM_LINK_TOKEN_MAP:
        return [ _COMM_LINK_TOKEN_MAP[lowered].value ]

    found: list[str] = []
    for part in re.split(r'[+/,;]+', text):
        token = part.strip().lower()
        if not token:
            continue
        choice = _COMM_LINK_TOKEN_MAP.get(token)
        if choice:
            found.append(choice.value)
            continue
        if 'старлинк' in token or 'starlink' in token:
            found.append(OperatorCommLink.STARLINK.value)
        elif 'бшпд' in token:
            found.append(OperatorCommLink.BSHPD.value)
        elif 'оптика' in token:
            found.append(OperatorCommLink.OPTICS.value)
        elif 'скс' in token:
            found.append(OperatorCommLink.SKS.value)
        elif token == 'lte':
            found.append(OperatorCommLink.LTE.value)
        elif token == 'rj45':
            found.append(OperatorCommLink.RJ45.value)
        elif '274' in token:
            found.append(OperatorCommLink.P274M.value)
    return normalize_comm_links(found)


def _normalize_text(value) -> str:
    if value is None:
        return ''
    text = str(value).replace('\n', ' ').strip()
    return re.sub(r'\s+', ' ', text)


def _parse_shift(value) -> str | None:
    text = _normalize_text(value).lower()
    if not text:
        return None
    return _SHIFT_MAP.get(text)


def _parse_time_value(value) -> time | None:
    if value is None or value == '':
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    text = _normalize_text(value)
    if not text:
        return None
    match = _TIME_RANGE_RE.search(text)
    if not match:
        for fmt in ('%H:%M', '%H.%M', '%H:%M:%S'):
            try:
                return datetime.strptime(text, fmt).time()
            except ValueError:
                continue
        return None
    hour, minute, _, _ = match.groups()
    return time(int(hour), int(minute))


def _parse_duty_range(value) -> tuple[time | None, time | None]:
    if value is None or value == '':
        return None, None
    if isinstance(value, time):
        return value, None
    if isinstance(value, datetime):
        return value.time(), None
    text = _normalize_text(value)
    if not text:
        return None, None
    match = _TIME_RANGE_RE.search(text)
    if not match:
        single = _parse_time_value(text)
        return single, None
    h1, m1, h2, m2 = match.groups()
    return time(int(h1), int(m1)), time(int(h2), int(m2))


def read_raschety_rows(file_path: str, *, sheet_name: str = SHEET_NAME) -> list[RaschetyRow]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f'Лист «{sheet_name}» не найден. Доступны: {", ".join(wb.sheetnames)}')
    ws = wb[sheet_name]
    rows: list[RaschetyRow] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        callname = _normalize_text(row[0] if len(row) > 0 else None)
        if not callname or callname.upper() in SKIP_OPERATOR_NAMES:
            continue
        engineer = _normalize_text(row[2] if len(row) > 2 else None)
        placement_zone = _parse_shift(row[4] if len(row) > 4 else None)
        duty_start, duty_end = _parse_duty_range(row[5] if len(row) > 5 else None)
        settlement = _normalize_text(row[8] if len(row) > 8 else None)
        location_name = _normalize_text(row[9] if len(row) > 9 else None)
        comm_link_raw = _normalize_text(row[7] if len(row) > 7 else None)
        drone_raw = _normalize_text(row[6] if len(row) > 6 else None)
        group_name = _normalize_text(row[11] if len(row) > 11 else None) or None
        rows.append(
            RaschetyRow(
                callname=callname,
                engineer_callname=engineer,
                placement_zone=placement_zone,
                duty_start=duty_start,
                duty_end=duty_end,
                comm_link_raw=comm_link_raw,
                drone_raw=drone_raw,
                settlement=settlement,
                location_name=location_name,
                group_name=group_name,
            )
        )
    wb.close()
    return rows


def collect_locations_from_rows(rows: list[RaschetyRow]) -> list[RaschetyLocation]:
    """Уникальные расположения: название (J), н.п. (I), связь (H)."""
    by_name: dict[str, dict[str, object]] = {}
    for row in rows:
        if not row.location_name:
            continue
        bucket = by_name.setdefault(row.location_name, {'settlement': '', 'comm_links': []})
        settlement = bucket['settlement']
        if row.settlement and (not settlement or settlement == row.settlement):
            bucket['settlement'] = row.settlement
        links = parse_comm_links_from_cell(row.comm_link_raw)
        for link in links:
            if link not in bucket['comm_links']:
                bucket['comm_links'].append(link)
    return [
        RaschetyLocation(
            name=name,
            settlement=str(data['settlement']),
            comm_links=tuple(data['comm_links']),
        )
        for name, data in sorted(by_name.items(), key=lambda item: item[0].lower())
    ]


def replace_locations_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyLocation], ImportStats]:
    rows = read_raschety_rows(file_path, sheet_name=sheet_name)
    locations = collect_locations_from_rows(rows)
    stats = ImportStats(rows_read=len(rows), locations_created=len(locations))

    ctx = transaction.atomic()
    with ctx:
        stats.locations_deleted = OperatorLocation.objects.count()
        if dry_run:
            transaction.set_rollback(True)
            return locations, stats

        OperatorProfile.objects.update(location_id=None, senior_id=None)
        OperatorLocation.objects.all().delete()
        for index, item in enumerate(locations):
            OperatorLocation.objects.create(
                name=item.name,
                description=item.settlement,
                comm_links=list(item.comm_links),
                sort_order=(index + 1) * 10,
                is_active=True,
            )

    return locations, stats


def update_locations_comm_links_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyLocation], ImportStats]:
    rows = read_raschety_rows(file_path, sheet_name=sheet_name)
    locations = collect_locations_from_rows(rows)
    stats = ImportStats(rows_read=len(rows))

    ctx = transaction.atomic()
    with ctx:
        for item in locations:
            if not item.comm_links:
                continue
            location = OperatorLocation.objects.filter(name=item.name).first()
            if not location:
                stats.skipped += 1
                continue
            if normalize_comm_links(location.comm_links) == list(item.comm_links):
                continue
            if not dry_run:
                location.comm_links = list(item.comm_links)
                location.save(update_fields=['comm_links', 'modified'])
            stats.locations_updated += 1

        if dry_run:
            transaction.set_rollback(True)

    return locations, stats


def _ig_groups_from_rows(rows: list[RaschetyRow]) -> list[str]:
    return sorted({row.group_name for row in rows if row.group_name})


def assign_pilot_groups_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyRow], ImportStats]:
    """Назначить пользователям Django-группы из столбца L (без изменения старшего оператора)."""
    rows = read_raschety_rows(file_path, sheet_name=sheet_name)
    stats = ImportStats(rows_read=len(rows))
    group_names = _ig_groups_from_rows(rows)
    if not group_names:
        return rows, stats

    django_groups = {
        group.name: group
        for group in Group.objects.filter(name__in=group_names)
    }
    missing_groups = [name for name in group_names if name not in django_groups]
    if missing_groups:
        raise ValueError(
            f'В админке нет групп: {", ".join(missing_groups)}. Создайте их вручную.'
        )

    ig_group_ids = [group.pk for group in django_groups.values()]

    ctx = transaction.atomic()
    with ctx:
        for row in rows:
            if not row.group_name:
                continue
            pilot = _find_pilot(row.callname)
            if not pilot:
                stats.skipped += 1
                continue

            user = User.objects.filter(pilot=pilot).first()
            if not user:
                stats.skipped += 1
                continue

            if dry_run:
                stats.groups_assigned += 1
                continue

            user.groups.remove(*Group.objects.filter(pk__in=ig_group_ids))
            user.groups.add(django_groups[row.group_name])
            stats.users_in_groups += 1
            stats.groups_assigned += 1

        if dry_run:
            transaction.set_rollback(True)

    return rows, stats


def assign_pilot_locations_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyRow], ImportStats]:
    rows = read_raschety_rows(file_path, sheet_name=sheet_name)
    stats = ImportStats(rows_read=len(rows))
    location_cache: dict[str, OperatorLocation | None] = {}

    ctx = transaction.atomic()
    with ctx:
        stats.pilot_locations_cleared = OperatorProfile.objects.exclude(location_id__isnull=True).count()
        if not dry_run:
            OperatorProfile.objects.update(location_id=None)

        for row in rows:
            if not row.location_name:
                continue

            if row.location_name not in location_cache:
                location_cache[row.location_name] = OperatorLocation.objects.filter(
                    name=row.location_name,
                    is_active=True,
                ).first()

            location = location_cache[row.location_name]
            if not location:
                stats.skipped += 1
                continue

            pilot = _find_pilot(row.callname)
            if not pilot:
                stats.skipped += 1
                continue

            profile = OperatorProfile.objects.filter(pilot=pilot).first()
            if not profile:
                stats.skipped += 1
                continue

            if dry_run:
                stats.pilot_locations_assigned += 1
                continue

            if profile.location_id != location.pk:
                profile.location = location
                profile.save(update_fields=['location', 'modified'])
            stats.pilot_locations_assigned += 1

        if dry_run:
            transaction.set_rollback(True)

    return rows, stats


def _drone_match_key(value: str) -> str:
    text = _normalize_text(value).lower().replace('«', '"').replace('»', '"')
    for prefix in ('fpv-дрон ', 'fpv-dron '):
        if text.startswith(prefix):
            text = text[len(prefix):]
    text = text.strip(' "\'')
    return re.sub(r'[\s\-–—/]', '', text)


def _build_drone_lookup() -> dict[str, Drone]:
    lookup: dict[str, Drone] = {}
    for drone in Drone.objects.all():
        for candidate in (drone.name,):
            key = _drone_match_key(candidate)
            if key and key not in lookup:
                lookup[key] = drone
    return lookup


def resolve_drone_from_cell(value: str, *, lookup: dict[str, Drone] | None = None) -> Drone | None:
    raw = _normalize_text(value)
    if not raw:
        return None
    lookup = lookup if lookup is not None else _build_drone_lookup()

    for candidate in (raw, raw.split(' DJI ')[0].strip(), raw.split(' FPV')[0].strip()):
        if not candidate:
            continue
        drone = Drone.objects.filter(name=candidate).first()
        if drone:
            return drone
        drone = Drone.objects.filter(name__iexact=candidate).first()
        if drone:
            return drone
        key = _drone_match_key(candidate)
        if key in lookup:
            return lookup[key]
    return None


def _find_pilot(callname: str) -> Pilot | None:
    pilot = Pilot.objects.filter(callname__iexact=callname).order_by('callname').first()
    if pilot:
        return pilot
    return Pilot.objects.filter(callname=callname).first()


def _apply_shift_times(profile: OperatorProfile, row: RaschetyRow) -> None:
    if row.placement_zone == OperatorPlacementZone.DAY:
        profile.day_shift_start = DAY_SHIFT_START
        profile.day_shift_end = DAY_SHIFT_END
        profile.night_shift_start = None
        profile.night_shift_end = None
    elif row.placement_zone == OperatorPlacementZone.NIGHT:
        profile.night_shift_start = NIGHT_SHIFT_START
        profile.night_shift_end = NIGHT_SHIFT_END
        profile.day_shift_start = None
        profile.day_shift_end = None


def sync_duty_roster_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyRow], ImportStats]:
    rows = read_raschety_rows(file_path, sheet_name=sheet_name)
    stats = ImportStats(rows_read=len(rows))
    drone_lookup = _build_drone_lookup()

    ctx = transaction.atomic()
    with ctx:
        stats.profiles_deactivated = OperatorProfile.objects.filter(is_active=True).count()
        if not dry_run:
            OperatorProfile.objects.filter(is_active=True).update(is_active=False)

        for row in rows:
            pilot = _find_pilot(row.callname)
            created_pilot = False
            pilot_fields: list[str] = []
            drone = resolve_drone_from_cell(row.drone_raw, lookup=drone_lookup) if row.drone_raw else None
            if row.drone_raw:
                if drone:
                    stats.drones_matched += 1
                else:
                    stats.drones_skipped += 1

            if pilot is None:
                if dry_run:
                    stats.pilots_created += 1
                    stats.profiles_created += 1
                    continue
                pilot = Pilot.objects.create(
                    callname=row.callname,
                    engineer_callname=row.engineer_callname or None,
                    drone_type=drone.name if drone else None,
                    tg_id=None,
                )
                created_pilot = True
                stats.pilots_created += 1
            else:
                new_engineer = row.engineer_callname or None
                if pilot.engineer_callname != new_engineer:
                    pilot.engineer_callname = new_engineer
                    pilot_fields.append('engineer_callname')
                if drone and pilot.drone_type != drone.name:
                    pilot.drone_type = drone.name
                    pilot_fields.append('drone_type')
                elif not drone and row.drone_raw:
                    pass
                if pilot_fields:
                    if not dry_run:
                        pilot_fields.append('modified')
                        pilot.save(update_fields=pilot_fields)
                    stats.pilots_updated += 1

            if dry_run:
                stats.profiles_created += 1
                continue

            profile, profile_created = OperatorProfile.objects.get_or_create(
                pilot=pilot,
                defaults={
                    'is_active': True,
                    'placement_zone': row.placement_zone or OperatorPlacementZone.DAY,
                },
            )
            if profile_created:
                stats.profiles_created += 1
            else:
                stats.profiles_updated += 1

            profile_changed = profile_created or not profile.is_active
            profile.is_active = True

            if row.placement_zone:
                if profile.placement_zone != row.placement_zone:
                    profile.placement_zone = row.placement_zone
                    profile_changed = True
                before = (
                    profile.day_shift_start,
                    profile.day_shift_end,
                    profile.night_shift_start,
                    profile.night_shift_end,
                )
                _apply_shift_times(profile, row)
                after = (
                    profile.day_shift_start,
                    profile.day_shift_end,
                    profile.night_shift_start,
                    profile.night_shift_end,
                )
                if before != after:
                    profile_changed = True

            if profile_changed or created_pilot:
                profile.save()

        if dry_run:
            transaction.set_rollback(True)

    return rows, stats


def import_pilots_from_raschety(
    file_path: str,
    *,
    sheet_name: str = SHEET_NAME,
    dry_run: bool = False,
) -> tuple[list[RaschetyRow], ImportStats]:
    return sync_duty_roster_from_raschety(
        file_path,
        sheet_name=sheet_name,
        dry_run=dry_run,
    )
