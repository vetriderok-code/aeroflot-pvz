"""Пятничный отчёт: лист «Цели», сводка «Итог», выгрузка Excel и Word."""

from __future__ import annotations

import io
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from xml.sax.saxutils import escape
from zoneinfo import ZoneInfo

from django.conf import settings
from django.db.models import Q
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from flights.api.reports import normalize_target_name
from flights.models import Flight
from flights.utils.nearest_settlement import resolve_settlements_batch

DIRECTION_EAST_THRESHOLD = 7_350_000
DIRECTION_NORTH_KHARKIV_THRESHOLD = 5_380_000
PROVISION_TARGET = 'обеспечение'


def squad_name() -> str:
    return settings.PORTAL_SITE_NAME.upper()


TARGET_ABBREVIATIONS = {
    'огневая позиция': 'ОП',
    'инженерное сооружение': 'ИС',
    'пункт временной дислокации': 'ПВД',
    'бпла': 'БпЛА',
    'автомобильная техника': 'АТ',
    'ббм': 'ББМ',
    'обеспечение': 'ОС',
    'наземные дроны': 'НД',
    'средства жизнеобеспечения': 'СЖО',
}

DIRECTIONS = (
    ('zaporizhzhia', 'Запорожское', 'Запорожском'),
    ('dobropillia', 'Добропольское', 'Добропольском'),
    ('kharkiv', 'Харьковское', 'Харьковском'),
)

MONTHS_GENITIVE = (
    '',
    'января',
    'февраля',
    'марта',
    'апреля',
    'мая',
    'июня',
    'июля',
    'августа',
    'сентября',
    'октября',
    'ноября',
    'декабря',
)

CELI_HEADERS = (
    '№',
    'Характер цели',
    'Дрон',
    'Координата X',
    'Координата Y',
    'Направление',
    'Ближайший н.п.',
)

DOCX_FONT = 'Times New Roman'
DOCX_FONT_SIZE = '28'  # half-points → 14 pt


@dataclass
class DocxRun:
    text: str
    bold: bool = False


@dataclass
class DocxParagraph:
    runs: list[DocxRun]
    align: str | None = None

    @classmethod
    def line(cls, text: str, *, bold: bool = False, align: str | None = None) -> DocxParagraph:
        return cls(runs=[DocxRun(text=text, bold=bold)], align=align)


@dataclass
class FridayReportSummary:
    period_from: date
    period_to: date
    destroyed_by_direction: dict[str, Counter[str]] = field(default_factory=dict)
    provision_own: dict[str, int] = field(default_factory=dict)
    provision_allied: dict[str, int] = field(default_factory=dict)
    provision_combined: dict[str, int] = field(default_factory=dict)
    total_destroyed: int = 0


def _msk_today(now=None) -> date:
    if isinstance(now, date) and not isinstance(now, datetime):
        return now
    now = now or timezone.now()
    if timezone.is_naive(now):
        from datetime import timezone as dt_timezone

        now = timezone.make_aware(now, dt_timezone.utc)
    return now.astimezone(ZoneInfo('Europe/Moscow')).date()


def friday_report_week_dates(*, week_offset: int = 0, now=None) -> tuple[date, date]:
    """Календарная рабочая неделя: понедельник–пятница."""
    today = _msk_today(now)
    this_monday = today - timedelta(days=today.weekday())
    monday = this_monday + timedelta(weeks=week_offset)
    friday = monday + timedelta(days=4)
    return monday, friday


def parse_report_period(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> tuple[date, date]:
    if date_from and date_to:
        start = date.fromisoformat(date_from)
        end = date.fromisoformat(date_to)
        if start > end:
            start, end = end, start
        return start, end
    offset = 0 if week_offset is None else week_offset
    return friday_report_week_dates(week_offset=offset, now=now)


def _format_target_name(value: str | None) -> str:
    return (value or '').strip().lower()


def _display_target_name(value: str) -> str:
    normalized = normalize_target_name(value)
    if normalized == 'бпла':
        return 'БпЛА'
    return normalized


def _is_combat_success_result(flight: Flight) -> bool:
    """Боевая цель для «Уничтожено»: в сводной — результат «поражено»."""
    result_raw = (flight.result_raw or '').strip().lower()
    if 'не усп' in result_raw or 'не пораж' in result_raw:
        return False
    return result_raw == 'поражено'


def _target_abbreviation(target: str) -> str:
    normalized = _format_target_name(target)
    if normalized in TARGET_ABBREVIATIONS:
        return TARGET_ABBREVIATIONS[normalized]
    display = _display_target_name(target)
    if display:
        return display
    return '—'


def _plural_objects(count: int) -> str:
    n = abs(count) % 100
    if 11 <= n <= 19:
        return 'объектов'
    remainder = n % 10
    if remainder == 1:
        return 'объект'
    if 2 <= remainder <= 4:
        return 'объекта'
    return 'объектов'


def _format_period_full(period_from: date, period_to: date) -> str:
    return (
        f'с {period_from.strftime("%d.%m.%Y")} по {period_to.strftime("%d.%m.%Y")} г.'
    )


def _group_rows_by_drone(rows: list[dict]) -> list[tuple[str, list[dict]]]:
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        drone = row.get('drone') or 'не указан'
        grouped.setdefault(drone, []).append(row)
    return sorted(
        grouped.items(),
        key=lambda item: (-len(item[1]), item[0].casefold()),
    )


def _format_drone_name(value: str | None) -> str:
    text = (value or '').strip()
    if not text:
        return ''
    return text.replace('V.', 'v.').replace('V ', 'v ')


def _resolve_sk42_meters(flight: Flight) -> tuple[int | None, int | None]:
    normalized = Flight.normalize_coordinates_field(flight.coordinates or '')
    if not normalized:
        return None, None
    parts = normalized.split()
    if len(parts) != 2:
        return None, None
    return int(parts[0]), int(parts[1])


def _direction_key_from_label(label: str) -> str | None:
    text = (label or '').strip().lower()
    if not text:
        return None
    if 'харьков' in text:
        return 'kharkiv'
    if 'доброполь' in text or 'добропил' in text:
        return 'dobropillia'
    if 'запорож' in text or 'зapor' in text:
        return 'zaporizhzhia'
    return None


def resolve_direction_key(
    *,
    north_meters: int | None,
    east_meters: int | None,
    flight: Flight,
) -> str:
    explicit = _direction_key_from_label(flight.direction or '')
    if explicit:
        return explicit

    if north_meters is not None and north_meters >= DIRECTION_NORTH_KHARKIV_THRESHOLD:
        return 'kharkiv'
    if east_meters is not None and east_meters >= DIRECTION_EAST_THRESHOLD:
        return 'dobropillia'
    return 'zaporizhzhia'


def resolve_direction_label(**kwargs) -> str:
    flight = kwargs.get('flight')
    if flight and (flight.direction or '').strip():
        explicit = _direction_key_from_label(flight.direction)
        if explicit is None:
            return flight.direction.strip()

    key = resolve_direction_key(**kwargs)
    for item_key, label, _ in DIRECTIONS:
        if item_key == key:
            return label
    return ''


def _is_provision_flight(flight: Flight) -> bool:
    """Обеспечено: характер цели «обеспечение»."""
    return _format_target_name(flight.target) == PROVISION_TARGET


def _provision_bucket(flight: Flight) -> str:
    text = ' '.join(
        part.strip().lower()
        for part in (flight.comment or '', flight.corrective or '')
        if part and part.strip()
    )
    if 'союз' in text:
        return 'allied'
    if 'свои' in text or 'собствен' in text:
        return 'own'
    return 'combined'


def _ensure_wgs84(flight: Flight) -> tuple[float | None, float | None]:
    lat = flight.lat_wgs84
    lon = flight.lon_wgs84
    if lat is None or lon is None:
        flight.get_coordinates_info_cached()
        lat = flight.lat_wgs84
        lon = flight.lon_wgs84
    return lat, lon


def get_friday_report_flights(*, period_from: date, period_to: date):
    return (
        Flight.objects.select_related('pilot')
        .filter(
            flight_date__gte=period_from,
            flight_date__lte=period_to,
        )
        .exclude(Q(coordinates__isnull=True) | Q(coordinates=''))
        .order_by('flight_date', 'flight_time', 'number', 'created')
    )


def _flight_rows(flights) -> list[dict]:
    rows: list[dict] = []
    for flight in flights:
        is_provision = _is_provision_flight(flight)
        if not is_provision and not _is_combat_success_result(flight):
            continue

        north, east = _resolve_sk42_meters(flight)
        if north is None or east is None:
            continue

        lat, lon = _ensure_wgs84(flight)
        direction_key = resolve_direction_key(
            north_meters=north,
            east_meters=east,
            flight=flight,
        )
        target = _format_target_name(flight.target)

        rows.append({
            'target': target,
            'target_display': _display_target_name(flight.target or ''),
            'drone': _format_drone_name(flight.drone),
            'coord_x': north,
            'coord_y': east,
            'direction_key': direction_key,
            'direction': resolve_direction_label(
                north_meters=north,
                east_meters=east,
                flight=flight,
            ),
            'lat': lat,
            'lon': lon,
            'is_provision': is_provision,
            'provision_bucket': _provision_bucket(flight),
        })
    return rows


def _build_summary(rows: list[dict], *, period_from: date, period_to: date) -> FridayReportSummary:
    summary = FridayReportSummary(period_from=period_from, period_to=period_to)
    for direction_key, _, _ in DIRECTIONS:
        summary.destroyed_by_direction[direction_key] = Counter()

    for row in rows:
        direction_key = row['direction_key']
        if row['is_provision']:
            summary.provision_allied[direction_key] = summary.provision_allied.get(direction_key, 0) + 1
            continue

        target_display = row['target_display']
        summary.destroyed_by_direction[direction_key][target_display] += 1
        summary.total_destroyed += 1

    return summary


def format_period_ru(period_from: date, period_to: date) -> str:
    if period_from.year != period_to.year:
        return (
            f'с {period_from.day} {MONTHS_GENITIVE[period_from.month]} {period_from.year} '
            f'по {period_to.day} {MONTHS_GENITIVE[period_to.month]} {period_to.year}'
        )
    return (
        f'с {period_from.day} {MONTHS_GENITIVE[period_from.month]} '
        f'по {period_to.day} {MONTHS_GENITIVE[period_to.month]}'
    )


def _plural_targets(count: int) -> str:
    n = abs(count) % 100
    if 11 <= n <= 19:
        return 'целей'
    remainder = n % 10
    if remainder == 1:
        return 'цель'
    if 2 <= remainder <= 4:
        return 'цели'
    return 'целей'


def _format_destroyed_line(target_name: str, count: int, *, last: bool = False) -> str:
    suffix = f' – {count:02d} ед.' if last else f' – {count:02d} ед.;'
    return f'{target_name}{suffix}'


def _format_provision_lines(summary: FridayReportSummary, direction_key: str) -> list[str]:
    own = summary.provision_own.get(direction_key, 0)
    allied = (
        summary.provision_allied.get(direction_key, 0)
        + summary.provision_combined.get(direction_key, 0)
    )

    own_line = f'свои силы – {own:02d} ед.' if own else 'свои силы –'
    allied_line = f'союзные силы - {allied:02d} ед.' if allied else 'союзные силы -'
    return [own_line, allied_line]


def build_summary_lines(summary: FridayReportSummary) -> list[str]:
    lines = [f'В период {format_period_ru(summary.period_from, summary.period_to)}']

    for direction_key, _direction_label, direction_prepositional in DIRECTIONS:
        lines.append(f'на {direction_prepositional} ТН:')
        lines.append('Уничтожено:')

        destroyed_items = sorted(
            summary.destroyed_by_direction.get(direction_key, Counter()).items(),
            key=lambda item: item[0].casefold(),
        )
        for index, (target_name, count) in enumerate(destroyed_items):
            lines.append(_format_destroyed_line(target_name, count, last=index == len(destroyed_items) - 1))

        lines.append('Обеспечено:')
        lines.extend(_format_provision_lines(summary, direction_key))

    lines.append(f'ВСЕГО УНИЧТОЖЕНО: {summary.total_destroyed} {_plural_targets(summary.total_destroyed)}.')
    return lines


def build_summary_docx_paragraphs(summary: FridayReportSummary) -> list[DocxParagraph]:
    paragraphs: list[DocxParagraph] = [
        DocxParagraph.line(
            f'В период {format_period_ru(summary.period_from, summary.period_to)}',
            align='both',
        ),
    ]

    for direction_key, _direction_label, direction_prepositional in DIRECTIONS:
        paragraphs.append(DocxParagraph.line(
            f'на {direction_prepositional} ТН:',
            bold=True,
            align='both',
        ))
        paragraphs.append(DocxParagraph.line('Уничтожено:', align='both'))

        destroyed_items = sorted(
            summary.destroyed_by_direction.get(direction_key, Counter()).items(),
            key=lambda item: item[0].casefold(),
        )
        for index, (target_name, count) in enumerate(destroyed_items):
            line = _format_destroyed_line(target_name, count, last=index == len(destroyed_items) - 1)
            align = 'both' if len(destroyed_items) == 1 else None
            paragraphs.append(DocxParagraph.line(line, align=align))

        paragraphs.append(DocxParagraph.line('Обеспечено:'))
        for line in _format_provision_lines(summary, direction_key):
            paragraphs.append(DocxParagraph.line(line))

    paragraphs.append(DocxParagraph(
        runs=[
            DocxRun('ВСЕГО УНИЧТОЖЕНО:', bold=True),
            DocxRun(f' {summary.total_destroyed} {_plural_targets(summary.total_destroyed)}.'),
        ],
    ))
    return paragraphs


def build_drone_docx_paragraphs(
    rows: list[dict],
    settlements: list[str],
    *,
    period_from: date,
    period_to: date,
) -> list[DocxParagraph]:
    """Раздел «ПРИМЕНЕНИЕ» по каждому дрону — как на 7-й странице образца."""
    enriched: list[dict] = []
    for row, settlement in zip(rows, settlements):
        enriched.append({
            **row,
            'settlement': settlement or '—',
        })

    paragraphs: list[DocxParagraph] = [DocxParagraph.line('')]

    for drone, drone_rows in _group_rows_by_drone(enriched):
        destroyed = [row for row in drone_rows if not row['is_provision']]
        provision_count = sum(1 for row in drone_rows if row['is_provision'])
        if not destroyed and not provision_count:
            continue

        paragraphs.append(DocxParagraph.line(f'ПРИМЕНЕНИЕ «{drone}»:'))

        period_prefix = (
            f'В период {_format_period_full(period_from, period_to)} '
            f'силами отряда «{squad_name()}»'
        )
        if destroyed and provision_count:
            stat_line = (
                f'{period_prefix} поражено {len(destroyed)} '
                f'{_plural_objects(len(destroyed))} противника, '
                f'доставлено {provision_count} ед. груза.'
            )
        elif provision_count:
            stat_line = f'{period_prefix} доставлено {provision_count} ед. груза.'
        else:
            stat_line = (
                f'{period_prefix} поражено {len(destroyed)} '
                f'{_plural_objects(len(destroyed))} противника.'
            )
        paragraphs.append(DocxParagraph.line(stat_line))

        if destroyed and not provision_count:
            paragraphs.append(DocxParagraph.line(
                f'В период {format_period_ru(period_from, period_to)} поражено '
                f'{len(destroyed)} {_plural_objects(len(destroyed))} противника.',
            ))

        if destroyed:
            paragraphs.append(DocxParagraph.line('Пораженные объекты противника:'))
            for index, row in enumerate(destroyed):
                abbr = _target_abbreviation(row['target'])
                line = (
                    f'{abbr} (X- {row["coord_x"]}; Y- {row["coord_y"]}) '
                    f'н.п. {row["settlement"]}'
                )
                line += '.' if index == len(destroyed) - 1 else ';'
                paragraphs.append(DocxParagraph.line(line))

    return paragraphs


def build_friday_report_docx_paragraphs(
    rows: list[dict],
    settlements: list[str],
    *,
    period_from: date,
    period_to: date,
) -> list[DocxParagraph]:
    summary = _build_summary(rows, period_from=period_from, period_to=period_to)
    paragraphs = build_summary_docx_paragraphs(summary)
    paragraphs.extend(build_drone_docx_paragraphs(
        rows,
        settlements,
        period_from=period_from,
        period_to=period_to,
    ))
    return paragraphs


def get_friday_report_preview(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> dict:
    period_from, period_to = parse_report_period(
        date_from=date_from,
        date_to=date_to,
        week_offset=week_offset,
        now=now,
    )
    flights = get_friday_report_flights(period_from=period_from, period_to=period_to)
    rows = _flight_rows(flights)
    summary = _build_summary(rows, period_from=period_from, period_to=period_to)

    by_target = Counter(
        row['target_display']
        for row in rows
        if row['target'] and not row['is_provision']
    )
    by_direction = Counter(row['direction'] for row in rows if row['direction'])

    return {
        'date_from': period_from.isoformat(),
        'date_to': period_to.isoformat(),
        'period_label': (
            f'{period_from.strftime("%d.%m.%Y")} – {period_to.strftime("%d.%m.%Y")} '
            f'(поражённые цели с координатами)'
        ),
        'total_targets': sum(1 for row in rows if not row['is_provision']),
        'total_destroyed': summary.total_destroyed,
        'summary_lines': build_summary_lines(summary),
        'targets': [
            {'name': name, 'count': count}
            for name, count in by_target.most_common(15)
        ],
        'directions': [
            {'name': name, 'count': count}
            for name, count in by_direction.most_common()
        ],
    }


def _write_celi_sheet(ws, rows: list[dict], settlements: list[str]) -> None:
    header_font = Font(bold=True, size=14)
    for col, title in enumerate(CELI_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.font = header_font

    widths = (4.5, 30.7, 30.7, 19.1, 13.0, 18.1, 22.1)
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for idx, (row, settlement) in enumerate(
        ((row, settlement) for row, settlement in zip(rows, settlements) if not row['is_provision']),
        start=1,
    ):
        excel_row = idx + 1
        ws.cell(row=excel_row, column=1, value=idx)
        ws.cell(row=excel_row, column=2, value=row['target_display'])
        ws.cell(row=excel_row, column=3, value=row['drone'])
        ws.cell(row=excel_row, column=4, value=row['coord_x'])
        ws.cell(row=excel_row, column=5, value=row['coord_y'])
        ws.cell(row=excel_row, column=6, value=row['direction'])
        ws.cell(row=excel_row, column=7, value=settlement or '—')


def _write_itog_sheet(ws, summary_lines: list[str]) -> None:
    ws.column_dimensions['A'].width = 100
    for idx, line in enumerate(summary_lines, start=1):
        ws.cell(row=idx, column=1, value=line)


@dataclass
class FridayReportData:
    period_from: date
    period_to: date
    rows: list[dict]
    summary: FridayReportSummary
    summary_lines: list[str]
    settlements: list[str]


def _collect_friday_report_data(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> FridayReportData:
    period_from, period_to = parse_report_period(
        date_from=date_from,
        date_to=date_to,
        week_offset=week_offset,
        now=now,
    )
    flights = get_friday_report_flights(period_from=period_from, period_to=period_to)
    rows = _flight_rows(flights)
    summary = _build_summary(rows, period_from=period_from, period_to=period_to)
    summary_lines = build_summary_lines(summary)
    settlements = resolve_settlements_batch([
        {'lat': row['lat'], 'lon': row['lon']}
        for row in rows
    ])
    return FridayReportData(
        period_from=period_from,
        period_to=period_to,
        rows=rows,
        summary=summary,
        summary_lines=summary_lines,
        settlements=settlements,
    )


def _build_friday_report_excel_from_data(data: FridayReportData) -> tuple[io.BytesIO, str]:
    wb = Workbook()
    ws_celi = wb.active
    ws_celi.title = 'Цели'
    _write_celi_sheet(ws_celi, data.rows, data.settlements)

    ws_itog = wb.create_sheet('Итог')
    _write_itog_sheet(ws_itog, data.summary_lines)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = (
        f'pyatnichnyj_otchet_{data.period_from.strftime("%Y%m%d")}_'
        f'{data.period_to.strftime("%Y%m%d")}.xlsx'
    )
    return buf, filename


def _build_friday_report_word_from_data(data: FridayReportData) -> tuple[io.BytesIO, str]:
    docx_paragraphs = build_friday_report_docx_paragraphs(
        data.rows,
        data.settlements,
        period_from=data.period_from,
        period_to=data.period_to,
    )

    buf = io.BytesIO()
    _write_formatted_docx(buf, docx_paragraphs)
    buf.seek(0)
    filename = (
        f'pyatnichnyj_otchet_{data.period_from.strftime("%Y%m%d")}_'
        f'{data.period_to.strftime("%Y%m%d")}.docx'
    )
    return buf, filename


def build_friday_report_excel(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> tuple[io.BytesIO, str]:
    data = _collect_friday_report_data(
        date_from=date_from,
        date_to=date_to,
        week_offset=week_offset,
        now=now,
    )
    return _build_friday_report_excel_from_data(data)


def build_friday_report_word(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> tuple[io.BytesIO, str]:
    data = _collect_friday_report_data(
        date_from=date_from,
        date_to=date_to,
        week_offset=week_offset,
        now=now,
    )
    return _build_friday_report_word_from_data(data)


def build_friday_report_archive(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
    week_offset: int | None = None,
    now=None,
) -> tuple[io.BytesIO, str]:
    data = _collect_friday_report_data(
        date_from=date_from,
        date_to=date_to,
        week_offset=week_offset,
        now=now,
    )
    excel_buf, excel_name = _build_friday_report_excel_from_data(data)
    word_buf, word_name = _build_friday_report_word_from_data(data)

    archive = io.BytesIO()
    with zipfile.ZipFile(archive, 'w', compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(excel_name, excel_buf.getvalue())
        zf.writestr(word_name, word_buf.getvalue())
    archive.seek(0)
    filename = (
        f'pyatnichnyj_otchet_{data.period_from.strftime("%Y%m%d")}_'
        f'{data.period_to.strftime("%Y%m%d")}.zip'
    )
    return archive, filename


def _docx_run_xml(run: DocxRun) -> str:
    rpr = (
        f'<w:rPr>'
        f'<w:rFonts w:ascii="{DOCX_FONT}" w:hAnsi="{DOCX_FONT}" w:cs="{DOCX_FONT}"/>'
        f'<w:sz w:val="{DOCX_FONT_SIZE}"/><w:szCs w:val="{DOCX_FONT_SIZE}"/>'
    )
    if run.bold:
        rpr += '<w:b/><w:bCs/>'
    rpr += '</w:rPr>'
    return f'<w:r>{rpr}<w:t xml:space="preserve">{escape(run.text)}</w:t></w:r>'


def _docx_paragraph_xml(paragraph: DocxParagraph) -> str:
    ppr = ''
    if paragraph.align:
        ppr = f'<w:pPr><w:jc w:val="{paragraph.align}"/></w:pPr>'
    runs = ''.join(_docx_run_xml(run) for run in paragraph.runs)
    return f'<w:p>{ppr}{runs}</w:p>'


def _write_formatted_docx(buffer: io.BytesIO, paragraphs: list[DocxParagraph]) -> None:
    body = ''.join(_docx_paragraph_xml(paragraph) for paragraph in paragraphs)
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f'<w:body>{body}<w:sectPr/></w:body>'
        '</w:document>'
    )
    content_types = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
        '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
        '<Default Extension="xml" ContentType="application/xml"/>'
        '<Override PartName="/word/document.xml" '
        'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
        '</Types>'
    )
    rels = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" '
        'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
        'Target="word/document.xml"/>'
        '</Relationships>'
    )

    with zipfile.ZipFile(buffer, 'w', compression=zipfile.ZIP_DEFLATED) as docx:
        docx.writestr('[Content_Types].xml', content_types)
        docx.writestr('_rels/.rels', rels)
        docx.writestr('word/document.xml', document_xml)


# Совместимость с прежними тестами/импортами
def _resolve_direction(*, east_meters: int | None, flight: Flight) -> str:
    return resolve_direction_label(
        north_meters=None,
        east_meters=east_meters,
        flight=flight,
    )
