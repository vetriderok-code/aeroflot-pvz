import asyncio
import io
import logging
import random

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm  # Для обычного входа
from django.http import FileResponse, JsonResponse
from django.shortcuts import render, redirect
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_http_methods
from django.utils.http import url_has_allowed_host_and_scheme

from flights.forms import TelegramAuthForm, TelegramCodeForm
from flights.models import User, Pilot
from telegram import Bot
from flights.utils.axes_logger import log_telegram_auth_attempt, log_telegram_code_attempt
from flights.utils.commander import get_post_login_url


BOT_TOKEN = settings.TOKEN

logger = logging.getLogger(__name__)


def resolve_login_redirect(request, user):
    next_url = request.POST.get('next') or request.GET.get('next')
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return get_post_login_url(user)


async def send_telegram_code(telegram_id, code):
    """Отправка кода в Telegram"""
    try:
        bot = Bot(token=BOT_TOKEN)
        await bot.send_message(
            chat_id=telegram_id,
            text=f"Ваш код подтверждения для входа: {code}\n\nКод действителен 5 минут."
        )
        return True
    except Exception as e:
        print(f"Ошибка отправки сообщения: {e}")
        return False


@require_http_methods(["GET", "POST"])
@csrf_protect
def login_view(request):
    """Универсальная страница входа - перенаправляет на стандартный вход"""
    if request.user.is_authenticated:
        return redirect(resolve_login_redirect(request, request.user))

    if request.method == 'POST':
        if 'login_type' in request.POST and request.POST['login_type'] == 'telegram':
            return telegram_login_step1_post(request)
        else:
            return standard_login_post(request)

    # Перенаправляем сразу на стандартный вход
    return redirect('standard_login')


@require_http_methods(["GET", "POST"])
@csrf_protect
def standard_login_view(request):
    """Обычный вход по username/password"""
    if request.user.is_authenticated:
        return redirect(resolve_login_redirect(request, request.user))

    if request.method == 'POST':
        return standard_login_post(request)

    form = AuthenticationForm()
    return render(request, 'login/login_standard.html', {
        'form': form,
        'next': request.GET.get('next', ''),
    })


def standard_login_post(request):
    """Обычный вход по username/password"""
    form = AuthenticationForm(request, data=request.POST)
    if form.is_valid():
        username = form.cleaned_data.get('username')
        password = form.cleaned_data.get('password')
        
        # Убрали проверку - пользователь может использовать любое имя, включая rubicon-app
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            messages.success(request, f"Добро пожаловать, {username}!")
            return redirect(resolve_login_redirect(request, user))
        else:
            messages.error(request, "Неверное имя пользователя или пароль. Убедитесь, что вы используете имя пользователя из Keycloak, а не название клиента.")
    else:
        # Показываем конкретные ошибки формы (но не дублируем non_field_errors, они уже показываются в шаблоне)
        if form.errors:
            for field, errors in form.errors.items():
                if field != '__all__':  # non_field_errors уже показываются в шаблоне
                    for error in errors:
                        messages.error(request, f"{field}: {error}")

    return render(request, 'login/login_standard.html', {
        'form': form,
        'next': request.POST.get('next', request.GET.get('next', '')),
    })


def telegram_login_step1_post(request):
    """Telegram вход - шаг 1"""
    username = request.POST.get('username')
    if not username:
        messages.error(request, "Пожалуйста, введите username")
        return render(request, 'login/login_telegram.html')

    try:
        user = User.objects.get(username=username)

        # Проверяем, есть ли у пользователя связанный пилот
        if not user.pilot:
            log_telegram_auth_attempt(request, username, successful=False)
            messages.error(request, "Для этого пользователя доступен только вход по паролю")
            return render(request, 'login/login_choice.html')

        # Генерируем код
        code = str(random.randint(100000, 999999))
        request.session['auth_code'] = code
        request.session['auth_username'] = username
        request.session.set_expiry(300)  # 5 минут

        # Отправляем код в Telegram
        success = asyncio.run(send_telegram_code(user.pilot.tg_id, code))

        if success:
            messages.success(request, "Код отправлен в ваш Telegram. Введите его ниже.")
            log_telegram_auth_attempt(request, username, successful=True)
            return redirect('telegram_login_step2')
        else:
            messages.error(request, "Не удалось отправить код. Попробуйте позже.")
            log_telegram_auth_attempt(request, username, successful=False)
            return render(request, 'login/login_telegram.html')

    except User.DoesNotExist:
        log_telegram_auth_attempt(request, username, successful=False)
        messages.error(request, "Пользователь не найден")
        return render(request, 'login/login_telegram.html')
    except Exception as e:
        logger.error(f"Ошибка при входе через Telegram: {e}")
        messages.error(request, "Произошла ошибка. Попробуйте позже.")
        return render(request, 'login/login_telegram.html')


@require_http_methods(["GET", "POST"])
@csrf_protect
def telegram_login_step1(request):
    """Telegram вход - шаг 1 (GET)"""
    if request.user.is_authenticated:
        return redirect('/')

    form = TelegramAuthForm()
    return render(request, 'login/login_telegram.html', {'form': form})


@require_http_methods(["GET", "POST"])
@csrf_protect
def telegram_login_step2(request):
    """Telegram вход - шаг 2 (ввод кода)"""
    if request.user.is_authenticated:
        return redirect('/')

    if 'auth_code' not in request.session or 'auth_username' not in request.session:
        messages.error(request, "Сессия истекла. Начните заново.")
        return redirect('telegram_login_step1')

    if request.method == 'POST':
        form = TelegramCodeForm(request.POST)
        if form.is_valid():
            code = form.cleaned_data.get('code')
            stored_code = request.session.get('auth_code')
            username = request.session.get('auth_username')

            if code == stored_code:
                try:
                    user = User.objects.get(username=username)
                    login(request, user)
                    log_telegram_code_attempt(request, username, successful=True)
                    messages.success(request, f"Добро пожаловать, {username}!")
                    # Очищаем сессию
                    del request.session['auth_code']
                    del request.session['auth_username']
                    return redirect(resolve_login_redirect(request, user))
                except User.DoesNotExist:
                    log_telegram_code_attempt(request, username, successful=False)
                    messages.error(request, "Пользователь не найден")
                    return redirect('telegram_login_step1')
            else:
                log_telegram_code_attempt(request, username, successful=False)
                messages.error(request, "Неверный код")
    else:
        form = TelegramCodeForm()

    return render(request, 'login/telegram_code.html', {'form': form})


@require_http_methods(["GET"])
def telegram_login_cancel(request):
    """Отмена входа через Telegram"""
    if 'auth_code' in request.session:
        del request.session['auth_code']
    if 'auth_username' in request.session:
        del request.session['auth_username']
    messages.info(request, "Вход отменен")
    return redirect('login')


@login_required(login_url='login')
def logout_view(request):
    """Выход из системы"""
    from django.contrib.auth import logout
    logout(request)
    messages.success(request, "Вы успешно вышли из системы")
    return redirect('login')


@login_required(login_url='login')
def debug_ip(request):
    """Отладочная функция для проверки IP"""
    ip = request.META.get('REMOTE_ADDR')
    forwarded = request.META.get('HTTP_X_FORWARDED_FOR')
    return JsonResponse({
        'ip': ip,
        'forwarded': forwarded,
        'user': str(request.user) if request.user.is_authenticated else 'anonymous'
    })


@login_required(login_url='login')
def map_view(request):
    context = {
        'yandex_api_key': settings.YANDEX_API_KEY,
        'map_ggc_enabled': settings.MAP_GSH_ENABLED,
        'map_ggc_tile_url': settings.MAP_GSH_BROWSER_TILE_URL,
        'map_ggc_zoom_min': settings.MAP_GSH_ZOOM_MIN,
        'map_ggc_zoom_max': settings.MAP_GSH_ZOOM_MAX,
    }
    response = render(request, 'map.html', context)
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response

@login_required(login_url='login')
def dashboard_view(request):
    from flights.utils.portal_features import require_dashboard_enabled
    require_dashboard_enabled()
    response = render(request, 'dashboard.html', {
        'yandex_api_key': settings.YANDEX_API_KEY,
    })
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response


@never_cache
@login_required(login_url='login')
def operator_dashboard_view(request):
    response = render(request, 'operator_dashboard.html')
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


@login_required(login_url='login')
def statistics_view(request):
    return render(request, 'statistics.html')

@login_required(login_url='login')
def schedule_view(request):
    return redirect('operator-dashboard-page')

@login_required(login_url='login')
def rating_view(request):
    return render(request, 'rating.html')

@login_required(login_url='login')
def reports_view(request):
    """Каталог отчётов"""
    return render(request, 'reports.html')


@login_required(login_url='login')
def friday_report_view(request):
    """Пятничный отчёт: выбор периода и превью."""
    from flights.utils.friday_report import friday_report_week_dates, get_friday_report_preview

    default_from, default_to = friday_report_week_dates(week_offset=0)
    date_from = request.GET.get('date_from') or default_from.isoformat()
    date_to = request.GET.get('date_to') or default_to.isoformat()

    try:
        preview = get_friday_report_preview(date_from=date_from, date_to=date_to)
    except ValueError:
        date_from = default_from.isoformat()
        date_to = default_to.isoformat()
        preview = get_friday_report_preview(date_from=date_from, date_to=date_to)

    return render(request, 'reports_friday.html', {
        'date_from': date_from,
        'date_to': date_to,
        'preview': preview,
    })


@login_required(login_url='login')
def friday_report_export_view(request):
    """Выгрузка пятничного отчёта: excel, word или zip."""
    from flights.utils.friday_report import (
        build_friday_report_archive,
        build_friday_report_excel,
        build_friday_report_word,
        friday_report_week_dates,
    )

    default_from, default_to = friday_report_week_dates(week_offset=0)
    date_from = request.GET.get('date_from') or default_from.isoformat()
    date_to = request.GET.get('date_to') or default_to.isoformat()
    export_kind = (request.GET.get('kind') or 'zip').lower()

    try:
        if export_kind == 'excel':
            buf, filename = build_friday_report_excel(date_from=date_from, date_to=date_to)
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif export_kind == 'word':
            buf, filename = build_friday_report_word(date_from=date_from, date_to=date_to)
            content_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        else:
            buf, filename = build_friday_report_archive(date_from=date_from, date_to=date_to)
            content_type = 'application/zip'
    except ValueError:
        return FileResponse(
            io.BytesIO('Некорректный период'.encode('utf-8')),
            as_attachment=True,
            filename='error.txt',
            content_type='text/plain; charset=utf-8',
            status=400,
        )

    return FileResponse(
        buf,
        as_attachment=True,
        filename=filename,
        content_type=content_type,
    )


@never_cache
@csrf_protect
@require_http_methods(['GET', 'POST'])
def tools_report_service_view(request):
    """Публичный сервис: загрузка Excel, обработка, ссылки на скачивание."""
    if request.method == 'POST':
        uploaded = request.FILES.get('excel_file')
        if not uploaded:
            messages.error(request, 'Выберите файл Excel (.xlsx).')
            return render(request, 'tools_report_service.html', {
                'north_col': request.POST.get('north_col', 'H'),
                'east_col': request.POST.get('east_col', 'G'),
                'settlement_col': request.POST.get('settlement_col', 'I'),
            })

        try:
            from flights.utils.excel_settlement_report import (
                ReportColumnConfig,
                process_report_full,
            )
            from flights.utils.tools_report_store import create_token, save_report

            try:
                columns = ReportColumnConfig.from_letters(
                    north=request.POST.get('north_col', 'H'),
                    east=request.POST.get('east_col', 'G'),
                    settlement=request.POST.get('settlement_col', 'I'),
                )
            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, 'tools_report_service.html', {
                    'north_col': request.POST.get('north_col', 'H'),
                    'east_col': request.POST.get('east_col', 'G'),
                    'settlement_col': request.POST.get('settlement_col', 'I'),
                })

            excel_bytes, kml_bytes, meta = process_report_full(uploaded, columns=columns)
            token = create_token()
            save_report(
                token,
                excel_bytes,
                meta.excel_filename,
                kml_bytes,
                meta.kml_filename,
                {
                    'source_name': meta.source_name,
                    'total_filled': meta.total_filled,
                    'point_count': meta.point_count,
                },
            )
            logger.info(
                'Tools report processed: file=%s filled=%s points=%s token=%s ip=%s',
                uploaded.name,
                meta.total_filled,
                meta.point_count,
                token[:8],
                request.META.get('REMOTE_ADDR'),
            )
            return redirect('tools-report-result', token=token)
        except ValueError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            logger.exception('Tools report service failed')
            messages.error(request, f'Ошибка обработки файла: {exc}')
        return render(request, 'tools_report_service.html', {
            'north_col': request.POST.get('north_col', 'H'),
            'east_col': request.POST.get('east_col', 'G'),
            'settlement_col': request.POST.get('settlement_col', 'I'),
        })

    return render(request, 'tools_report_service.html', {
        'north_col': 'H',
        'east_col': 'G',
        'settlement_col': 'I',
    })


@never_cache
@require_http_methods(['GET'])
def tools_report_result_view(request, token):
    from flights.utils.tools_report_store import get_meta

    meta = get_meta(token)
    if not meta:
        messages.error(request, 'Результат обработки не найден или истёк (хранится 1 час). Загрузите файл снова.')
        return redirect('tools-report-service')

    return render(request, 'tools_report_result.html', {
        'token': token,
        'meta': meta,
    })


@require_http_methods(['GET'])
def tools_report_download_view(request, token, kind):
    from flights.utils.tools_report_store import get_excel, get_kml

    if kind == 'excel':
        payload = get_excel(token)
        content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    elif kind == 'kml':
        payload = get_kml(token)
        content_type = 'application/vnd.google-earth.kml+xml'
    else:
        return redirect('tools-report-service')

    if not payload:
        messages.error(request, 'Файл не найден или срок хранения истёк.')
        return redirect('tools-report-service')

    data, filename = payload
    return FileResponse(
        io.BytesIO(data),
        as_attachment=True,
        filename=filename,
        content_type=content_type,
    )


@login_required(login_url='login')
def export_report_excel(request):
    """Экспорт отчета в Excel"""
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.drawing.image import Image
    import matplotlib
    matplotlib.use('Agg')  # Используем backend без GUI
    import matplotlib.pyplot as plt
    import io
    import base64
    from datetime import datetime
    from flights.api.reports import ReportsDataView
    from flights.models import Flight, FlightResultTypes
    
    try:
        # Получаем данные через API
        api_view = ReportsDataView()
        api_view.request = request
        response = api_view.get(request)
        
        if response.status_code != 200:
            return HttpResponse('Ошибка получения данных', status=500)
        
        data = response.data
        
        # Создаем рабочую книгу
        wb = Workbook()
        ws = wb.active
        ws.title = "Сводный отчет"
        
        # Стили
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Заголовок
        row = 1
        ws.merge_cells(f'A{row}:F{row}')
        title_cell = ws[f'A{row}']
        title_cell.value = f"Сводный отчет о полетах"
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2
        
        # Общая статистика
        ws[f'A{row}'] = "Общая статистика"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        stats_headers = ['Показатель', 'Значение']
        for col, header in enumerate(stats_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        stats_data = [
            ['Всего вылетов', data['summary']['total_flights']],
            ['Уничтожено', data['summary']['destroyed_flights']],
            ['Поражено', data['summary']['defeated_flights']],
            ['Не поражено', data['summary']['not_defeated_flights']],
            ['% Уничтожения', f"{data['summary']['destruction_rate_percent']}%"],
            ['% Успеха', f"{data['summary']['success_rate_percent']}%"],
        ]
        
        for stat_row in stats_data:
            for col, value in enumerate(stat_row, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = value
                cell.border = border
                if col == 2:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        row += 2
        
        # Статистика по пилотам
        ws[f'A{row}'] = "Статистика по пилотам"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        pilot_headers = ['Пилот', 'Всего', 'Уничтожено', 'Поражено', 'Не поражено', '% Успеха']
        for col, header in enumerate(pilot_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        pilot_start_row = row
        for pilot in data['pilots'][:20]:  # Ограничиваем до 20 пилотов
            ws.cell(row=row, column=1).value = pilot['pilot_name']
            ws.cell(row=row, column=2).value = pilot['total_flights']
            ws.cell(row=row, column=3).value = pilot['destroyed_flights']
            ws.cell(row=row, column=4).value = pilot['defeated_flights']
            ws.cell(row=row, column=5).value = pilot['not_defeated_flights']
            ws.cell(row=row, column=6).value = f"{pilot['success_rate_percent']}%"
            
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Диаграмма по пилотам
        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Топ-10 пилотов по количеству вылетов"
        chart.y_axis.title = 'Количество вылетов'
        chart.x_axis.title = 'Пилоты'
        
        data_ref = Reference(ws, min_col=2, min_row=pilot_start_row-1, max_row=min(row-1, pilot_start_row+9))
        cats_ref = Reference(ws, min_col=1, min_row=pilot_start_row, max_row=min(row-1, pilot_start_row+9))
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart.height = 10
        chart.width = 15
        
        ws.add_chart(chart, f'H{pilot_start_row}')
        
        row += 15
        
        # Статистика по целям
        ws[f'A{row}'] = "Статистика по целям"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        target_headers = ['Тип цели', 'Всего', 'Уничтожено', 'Поражено', 'Не поражено']
        for col, header in enumerate(target_headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        row += 1
        target_start_row = row
        for target in data['targets'][:15]:  # Ограничиваем до 15 целей
            ws.cell(row=row, column=1).value = target['target']
            ws.cell(row=row, column=2).value = target['total_flights']
            ws.cell(row=row, column=3).value = target['destroyed_flights']
            ws.cell(row=row, column=4).value = target['defeated_flights']
            ws.cell(row=row, column=5).value = target['not_defeated_flights']
            
            for col in range(1, 6):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                if col > 1:
                    cell.alignment = Alignment(horizontal='right')
            row += 1
        
        # Настройка ширины колонок
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # Создаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в Excel: {e}", exc_info=True)
        return HttpResponse(f'Ошибка: {str(e)}', status=500)


@login_required(login_url='login')
def export_report_pdf(request):
    """Экспорт отчета в PDF"""
    from django.http import HttpResponse
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    import io
    from datetime import datetime
    from flights.api.reports import ReportsDataView
    
    try:
        # Получаем данные через API
        api_view = ReportsDataView()
        api_view.request = request
        response = api_view.get(request)
        
        if response.status_code != 200:
            return HttpResponse('Ошибка получения данных', status=500)
        
        data = response.data
        
        # Создаем PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
        story = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#366092'),
            spaceAfter=12,
            spaceBefore=12
        )
        
        # Заголовок
        story.append(Paragraph("Сводный отчет о полетах", title_style))
        story.append(Spacer(1, 0.2*inch))
        
        # Общая статистика
        story.append(Paragraph("Общая статистика", heading_style))
        
        summary_data = [
            ['Показатель', 'Значение'],
            ['Всего вылетов', str(data['summary']['total_flights'])],
            ['Уничтожено', str(data['summary']['destroyed_flights'])],
            ['Поражено', str(data['summary']['defeated_flights'])],
            ['Не поражено', str(data['summary']['not_defeated_flights'])],
            ['% Уничтожения', f"{data['summary']['destruction_rate_percent']}%"],
            ['% Успеха', f"{data['summary']['success_rate_percent']}%"],
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Статистика по пилотам
        story.append(Paragraph("Статистика по пилотам (Топ-10)", heading_style))
        
        pilot_data = [['Пилот', 'Всего', 'Уничтожено', 'Поражено', '% Успеха']]
        for pilot in data['pilots'][:10]:
            pilot_data.append([
                pilot['pilot_name'],
                str(pilot['total_flights']),
                str(pilot['destroyed_flights']),
                str(pilot['defeated_flights']),
                f"{pilot['success_rate_percent']}%"
            ])
        
        pilot_table = Table(pilot_data, colWidths=[2*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        pilot_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(pilot_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Диаграмма по пилотам
        fig, ax = plt.subplots(figsize=(8, 5))
        top_pilots = data['pilots'][:10]
        pilot_names = [p['pilot_name'] for p in top_pilots]
        pilot_totals = [p['total_flights'] for p in top_pilots]
        
        ax.barh(pilot_names, pilot_totals, color='#366092')
        ax.set_xlabel('Количество вылетов')
        ax.set_title('Топ-10 пилотов по количеству вылетов')
        ax.invert_yaxis()
        plt.tight_layout()
        
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        
        img = Image(img_buffer, width=6*inch, height=3.75*inch)
        story.append(img)
        story.append(Spacer(1, 0.3*inch))
        
        # Статистика по целям
        story.append(Paragraph("Статистика по целям (Топ-10)", heading_style))
        
        target_data = [['Тип цели', 'Всего', 'Уничтожено', 'Поражено']]
        for target in data['targets'][:10]:
            target_data.append([
                target['target'],
                str(target['total_flights']),
                str(target['destroyed_flights']),
                str(target['defeated_flights'])
            ])
        
        target_table = Table(target_data, colWidths=[3*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        target_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        story.append(target_table)
        
        # Строим PDF
        doc.build(story)
        
        # Возвращаем ответ
        response = HttpResponse(content_type='application/pdf')
        filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(buffer.getvalue())
        buffer.close()
        return response
        
    except Exception as e:
        logger.error(f"Ошибка при экспорте в PDF: {e}", exc_info=True)
        return HttpResponse(f'Ошибка: {str(e)}', status=500)
