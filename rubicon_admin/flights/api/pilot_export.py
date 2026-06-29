from django.http import HttpResponse
from django.db.models import Count, Case, When, IntegerField, Sum
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from flights.models import Flight, FlightResultTypes, Pilot
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import logging

logger = logging.getLogger(__name__)


class PilotExportExcelView(APIView):
    def get(self, request, format=None):
        try:
            pilot_callname = request.query_params.get('pilot_callname')
            if not pilot_callname:
                return Response({'error': 'pilot_callname is required'}, status=status.HTTP_400_BAD_REQUEST)

            # Находим пилота
            pilot_callname_clean = pilot_callname.strip()
            try:
                pilot = Pilot.objects.get(callname__iexact=pilot_callname_clean)
            except Pilot.DoesNotExist:
                try:
                    pilot = Pilot.objects.get(callname=pilot_callname_clean)
                except Pilot.DoesNotExist:
                    logger.error(f"Пилот не найден: '{pilot_callname_clean}'")
                    return Response({'error': f'Pilot not found: {pilot_callname_clean}'}, status=status.HTTP_404_NOT_FOUND)
            except Pilot.MultipleObjectsReturned:
                pilot = Pilot.objects.filter(callname__iexact=pilot_callname_clean).first()
                logger.warning(f"Найдено несколько пилотов с именем '{pilot_callname_clean}', используется первый")

            # Получаем все полеты пилота
            flights = Flight.objects.filter(pilot=pilot).select_related('pilot').order_by('-flight_date', '-flight_time')

            # Создаем рабочую книгу
            wb = Workbook()
            
            # Удаляем лист по умолчанию, создадим свои
            wb.remove(wb.active)
            
            # Стили
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=18, color="1F4E78")
            subtitle_font = Font(bold=True, size=14, color="2E75B6")
            border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            left_alignment = Alignment(horizontal='left', vertical='center')
            
            # Цвета для результатов
            destroyed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            defeated_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            not_defeated_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Лист 1: Общая информация
            ws_summary = wb.create_sheet("Общая информация", 0)
            row = 1
            
            # Заголовок
            ws_summary.merge_cells(f'A{row}:E{row}')
            title_cell = ws_summary[f'A{row}']
            title_cell.value = f"Карточка пилота: {pilot.callname}"
            title_cell.font = title_font
            title_cell.alignment = center_alignment
            title_cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 2
            
            # Статистика
            total_flights = flights.count()
            destroyed_flights = flights.filter(result=FlightResultTypes.DESTROYED).count()
            defeated_flights = flights.filter(result=FlightResultTypes.DEFEATED).count()
            not_defeated_flights = flights.filter(result=FlightResultTypes.NOT_DEFEATED).count()
            destruction_rate = (destroyed_flights / total_flights * 100) if total_flights > 0 else 0
            success_rate = ((destroyed_flights + defeated_flights) / total_flights * 100) if total_flights > 0 else 0
            
            stats_data = [
                ['Показатель', 'Значение'],
                ['Всего вылетов', total_flights],
                ['Уничтожено', destroyed_flights],
                ['Поражено', defeated_flights],
                ['Не поражено', not_defeated_flights],
                ['% Уничтожения', f"{destruction_rate:.2f}%"],
                ['% Успеха', f"{success_rate:.2f}%"],
            ]
            
            for i, (label, value) in enumerate(stats_data):
                label_cell = ws_summary[f'A{row + i}']
                value_cell = ws_summary[f'B{row + i}']
                label_cell.value = label
                value_cell.value = value
                label_cell.font = Font(bold=True, size=11)
                value_cell.font = Font(size=11)
                label_cell.border = border
                value_cell.border = border
                label_cell.alignment = left_alignment
                value_cell.alignment = center_alignment
                if i == 0:  # Заголовок
                    label_cell.fill = header_fill
                    label_cell.font = header_font
                    value_cell.fill = header_fill
                    value_cell.font = header_font
                elif i == 2:  # Уничтожено
                    value_cell.fill = destroyed_fill
                elif i == 3:  # Поражено
                    value_cell.fill = defeated_fill
                elif i == 4:  # Не поражено
                    value_cell.fill = not_defeated_fill
            
            # Настройка ширины столбцов
            ws_summary.column_dimensions['A'].width = 20
            ws_summary.column_dimensions['B'].width = 15
            row += len(stats_data) + 2
            
            # Лист 2: Вылеты по целям
            ws_targets = wb.create_sheet("Вылеты по целям", 1)
            row = 1
            
            # Заголовок листа
            ws_targets.merge_cells(f'A{row}:F{row}')
            title_cell = ws_targets[f'A{row}']
            title_cell.value = f"Вылеты по целям - {pilot.callname}"
            title_cell.font = subtitle_font
            title_cell.alignment = center_alignment
            row += 2
            
            # Заголовки таблицы
            headers = ['Цель', 'Всего', 'Уничтожено', 'Поражено', 'Не поражено', '% Уничтожения']
            for col, header in enumerate(headers, 1):
                cell = ws_targets.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            row += 1
            
            # Данные по целям
            flights_by_target = flights.values('target').annotate(
                total=Count('id'),
                destroyed=Sum(Case(When(result=FlightResultTypes.DESTROYED, then=1), output_field=IntegerField())),
                defeated=Sum(Case(When(result=FlightResultTypes.DEFEATED, then=1), output_field=IntegerField())),
                not_defeated=Sum(Case(When(result=FlightResultTypes.NOT_DEFEATED, then=1), output_field=IntegerField())),
            ).order_by('-total')
            
            for target_stat in flights_by_target:
                target_name = target_stat['target'] or 'Без указания цели'
                destroyed = target_stat['destroyed'] if target_stat['destroyed'] is not None else 0
                defeated = target_stat['defeated'] if target_stat['defeated'] is not None else 0
                not_defeated = target_stat['not_defeated'] if target_stat['not_defeated'] is not None else 0
                total = target_stat['total']
                destruction_rate = (destroyed / total * 100) if total > 0 else 0
                
                data_row = [target_name, total, destroyed, defeated, not_defeated, f"{destruction_rate:.2f}%"]
                for col, value in enumerate(data_row, 1):
                    cell = ws_targets.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    cell.alignment = center_alignment if col > 1 else left_alignment
                    
                    if col == 3:  # Уничтожено
                        cell.fill = destroyed_fill
                    elif col == 4:  # Поражено
                        cell.fill = defeated_fill
                    elif col == 5:  # Не поражено
                        cell.fill = not_defeated_fill
                
                row += 1
            
            # Настройка ширины столбцов
            ws_targets.column_dimensions['A'].width = 30
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_targets.column_dimensions[col].width = 15
            
            # Лист 3: Детальный список вылетов
            ws_flights = wb.create_sheet("Детальный список", 2)
            row = 1
            
            # Заголовок листа
            ws_flights.merge_cells(f'A{row}:I{row}')
            title_cell = ws_flights[f'A{row}']
            title_cell.value = f"Детальный список вылетов - {pilot.callname}"
            title_cell.font = subtitle_font
            title_cell.alignment = center_alignment
            row += 2
            
            # Заголовки таблицы
            headers = ['№', 'Дата', 'Время', 'Цель', 'Дрон', 'Результат', 'Координаты', 'Дистанция', 'Комментарий']
            for col, header in enumerate(headers, 1):
                cell = ws_flights.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            
            row += 1
            
            # Данные вылетов (ограничиваем до 1000 записей для производительности)
            flights_list = flights[:1000]
            for flight in flights_list:
                result_text = ''
                result_fill = None
                if flight.result == FlightResultTypes.DESTROYED:
                    result_text = 'Уничтожен'
                    result_fill = destroyed_fill
                elif flight.result == FlightResultTypes.DEFEATED:
                    result_text = 'Поражен'
                    result_fill = defeated_fill
                elif flight.result == FlightResultTypes.NOT_DEFEATED:
                    result_text = 'Не поражен'
                    result_fill = not_defeated_fill
                
                date_str = flight.flight_date.strftime('%d.%m.%Y') if flight.flight_date else ''
                time_str = flight.flight_time.strftime('%H:%M') if flight.flight_time else ''
                
                data_row = [
                    flight.number or '',
                    date_str,
                    time_str,
                    flight.target or '',
                    flight.drone or '',
                    result_text,
                    flight.coordinates or '',
                    flight.distance or '',
                    flight.comment or '',
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = ws_flights.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border
                    if col == 6:  # Результат
                        cell.alignment = center_alignment
                        if result_fill:
                            cell.fill = result_fill
                    elif col in [1, 2, 3, 7, 8]:  # Номер, Дата, Время, Координаты, Дистанция
                        cell.alignment = center_alignment
                    else:
                        cell.alignment = left_alignment
                
                row += 1
            
            # Настройка ширины столбцов
            ws_flights.column_dimensions['A'].width = 8  # №
            ws_flights.column_dimensions['B'].width = 12  # Дата
            ws_flights.column_dimensions['C'].width = 10  # Время
            ws_flights.column_dimensions['D'].width = 25  # Цель
            ws_flights.column_dimensions['E'].width = 20  # Дрон
            ws_flights.column_dimensions['F'].width = 15  # Результат
            ws_flights.column_dimensions['G'].width = 20  # Координаты
            ws_flights.column_dimensions['H'].width = 12  # Дистанция
            ws_flights.column_dimensions['I'].width = 40  # Комментарий
            
            # Замораживаем первую строку с заголовками
            # Находим первую строку данных после заголовков
            targets_data_start = 4  # Заголовок листа (1) + пустая строка (2) + заголовки таблицы (3) + 1
            ws_targets.freeze_panes = f'A{targets_data_start}'
            ws_flights.freeze_panes = 'A4'  # Заголовок листа (1) + пустая строка (2) + заголовки таблицы (3) + 1
            
            # Лист 4: Использование БК (боеприпасы/компоненты)
            ws_bc = wb.create_sheet("Использование БК", 3)
            row = 1
            
            # Заголовок листа
            ws_bc.merge_cells(f'A{row}:B{row}')
            title_cell = ws_bc[f'A{row}']
            title_cell.value = f"Использование боеприпасов и компонентов - {pilot.callname}"
            title_cell.font = subtitle_font
            title_cell.alignment = center_alignment
            row += 2
            
            # Статистика по дронам
            ws_bc.merge_cells(f'A{row}:B{row}')
            section_title = ws_bc[f'A{row}']
            section_title.value = "Дроны"
            section_title.font = Font(bold=True, size=13, color="2E75B6")
            row += 1
            
            # Заголовки таблицы дронов
            headers = ['Дрон', 'Количество использований']
            for col, header in enumerate(headers, 1):
                cell = ws_bc.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            row += 1
            
            # Данные по дронам
            drone_usage = flights.exclude(drone__isnull=True).exclude(drone='').values('drone').annotate(
                count=Count('id')
            ).order_by('-count')
            
            for item in drone_usage:
                ws_bc.cell(row=row, column=1, value=item['drone']).border = border
                ws_bc.cell(row=row, column=1).alignment = left_alignment
                ws_bc.cell(row=row, column=2, value=item['count']).border = border
                ws_bc.cell(row=row, column=2).alignment = center_alignment
                row += 1
            
            row += 2
            
            # Статистика по боевой части (explosive_type)
            ws_bc.merge_cells(f'A{row}:B{row}')
            section_title = ws_bc[f'A{row}']
            section_title.value = "Боевая часть (ВВ)"
            section_title.font = Font(bold=True, size=13, color="2E75B6")
            row += 1
            
            # Заголовки таблицы боевой части
            headers = ['Тип ВВ', 'Количество использований']
            for col, header in enumerate(headers, 1):
                cell = ws_bc.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            row += 1
            
            # Данные по боевой части
            explosive_type_usage = flights.exclude(explosive_type__isnull=True).exclude(explosive_type='').values('explosive_type').annotate(
                count=Count('id')
            ).order_by('-count')
            
            for item in explosive_type_usage:
                ws_bc.cell(row=row, column=1, value=item['explosive_type']).border = border
                ws_bc.cell(row=row, column=1).alignment = left_alignment
                ws_bc.cell(row=row, column=2, value=item['count']).border = border
                ws_bc.cell(row=row, column=2).alignment = center_alignment
                row += 1
            
            row += 2
            
            # Статистика по запалам (explosive_device)
            ws_bc.merge_cells(f'A{row}:B{row}')
            section_title = ws_bc[f'A{row}']
            section_title.value = "Запалы (ВУ)"
            section_title.font = Font(bold=True, size=13, color="2E75B6")
            row += 1
            
            # Заголовки таблицы запалов
            headers = ['Тип ВУ', 'Количество использований']
            for col, header in enumerate(headers, 1):
                cell = ws_bc.cell(row=row, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border
            row += 1
            
            # Данные по запалам
            explosive_device_usage = flights.exclude(explosive_device__isnull=True).exclude(explosive_device='').values('explosive_device').annotate(
                count=Count('id')
            ).order_by('-count')
            
            for item in explosive_device_usage:
                ws_bc.cell(row=row, column=1, value=item['explosive_device']).border = border
                ws_bc.cell(row=row, column=1).alignment = left_alignment
                ws_bc.cell(row=row, column=2, value=item['count']).border = border
                ws_bc.cell(row=row, column=2).alignment = center_alignment
                row += 1
            
            # Настройка ширины столбцов для листа БК
            ws_bc.column_dimensions['A'].width = 40
            ws_bc.column_dimensions['B'].width = 25
            
            # Создаем ответ
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"pilot_{pilot.callname}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Ошибка при экспорте данных пилота в Excel: {e}", exc_info=True)
            return Response({'error': f'Error exporting to Excel: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

