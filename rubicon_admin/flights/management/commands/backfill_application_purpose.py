"""Заполняет «Цель применения» (колонка N) из сводной Excel."""

from __future__ import annotations

import datetime
import hashlib

from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from flights.models import Flight, Pilot
from flights.utils.excel_import_share import resolve_import_file


class Command(BaseCommand):
    help = 'Заполняет Flight.application_purpose из колонки N сводной Excel.'

    COL_TIME = 2
    COL_APPLICATION_PURPOSE = 14
    COL_FLIGHT_DATE = 18
    COL_FLIGHT_NUMBER = 19
    COL_OPERATOR_CALLNAME = 22
    BATCH_SIZE = 500

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            dest='file_path',
            help='Путь к сводной Excel (.xlsm/.xlsx). По умолчанию — файл из шары.',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Только показать, сколько записей будет обновлено.',
        )

    def handle(self, *args, **options):
        try:
            file_path = options.get('file_path') or str(resolve_import_file())
        except (FileNotFoundError, ValueError) as exc:
            raise CommandError(str(exc)) from exc

        pilots = {pilot.callname.lower(): pilot for pilot in Pilot.objects.all()}
        flight_index: dict[tuple, Flight] = {}
        for flight in Flight.objects.only(
            'id',
            'number',
            'pilot_id',
            'flight_date',
            'flight_time',
            'application_purpose',
        ).iterator(chunk_size=2000):
            flight_index[
                flight.number,
                flight.pilot_id,
                flight.flight_date,
                flight.flight_time,
            ] = flight

        pending: dict[tuple, str] = {}
        skipped = 0

        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        for row_idx in range(5, ws.max_row + 1):
            application_raw = ws.cell(row=row_idx, column=self.COL_APPLICATION_PURPOSE).value
            application_purpose = str(application_raw).strip() if application_raw else ''
            if not application_purpose:
                continue

            pilot_raw = ws.cell(row=row_idx, column=self.COL_OPERATOR_CALLNAME).value
            pilot_callname = str(pilot_raw).strip().lower() if pilot_raw else ''
            pilot = pilots.get(pilot_callname)
            if not pilot:
                skipped += 1
                continue

            date_value = ws.cell(row=row_idx, column=self.COL_FLIGHT_DATE).value
            if isinstance(date_value, datetime.datetime):
                flight_date = date_value.date()
            elif isinstance(date_value, datetime.date):
                flight_date = date_value
            else:
                skipped += 1
                continue

            flight_number_raw = ws.cell(row=row_idx, column=self.COL_FLIGHT_NUMBER).value
            if flight_number_raw is None or str(flight_number_raw).strip() == '':
                unique_str = f'{flight_date.isoformat()}_{pilot.id}_{row_idx}'
                hash_value = int(hashlib.md5(unique_str.encode()).hexdigest()[:8], 16)
                flight_number = abs(hash_value) % (10 ** 8)
            else:
                try:
                    flight_number = int(float(flight_number_raw))
                except (TypeError, ValueError):
                    skipped += 1
                    continue

            time_value = ws.cell(row=row_idx, column=self.COL_TIME).value
            if isinstance(time_value, datetime.datetime):
                flight_time = time_value.time()
            elif isinstance(time_value, datetime.time):
                flight_time = time_value
            elif time_value:
                try:
                    flight_time = datetime.datetime.strptime(str(time_value).strip(), '%H:%M:%S').time()
                except ValueError:
                    try:
                        flight_time = datetime.datetime.strptime(str(time_value).strip(), '%H:%M').time()
                    except ValueError:
                        skipped += 1
                        continue
            else:
                skipped += 1
                continue

            flight_key = (flight_number, pilot.id, flight_date, flight_time)
            flight = flight_index.get(flight_key)
            if not flight:
                skipped += 1
                continue

            if (flight.application_purpose or '') == application_purpose:
                continue

            pending[flight_key] = application_purpose

        wb.close()

        to_update: list[Flight] = []
        for flight_key, application_purpose in pending.items():
            flight = flight_index[flight_key]
            flight.application_purpose = application_purpose
            to_update.append(flight)

        if options['dry_run']:
            self.stdout.write(
                f'Будет обновлено записей: {len(to_update)}, пропущено: {skipped}'
            )
            return

        updated = 0
        for start in range(0, len(to_update), self.BATCH_SIZE):
            batch = to_update[start:start + self.BATCH_SIZE]
            Flight.objects.bulk_update(batch, ['application_purpose'], batch_size=self.BATCH_SIZE)
            updated += len(batch)
            self.stdout.write(f'Обновлено: {updated}/{len(to_update)}', ending='\r')

        self.stdout.write('')
        self.stdout.write(self.style.SUCCESS(
            f'Обновлено application_purpose у {updated} записей, пропущено: {skipped}.'
        ))
