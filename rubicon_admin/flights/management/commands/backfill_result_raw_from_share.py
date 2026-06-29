"""Синхронизирует result_raw и result из колонки O сводной Excel."""

from __future__ import annotations

import datetime
import hashlib

import openpyxl
from django.core.management.base import BaseCommand, CommandError

from flights.models import Flight, FlightResultTypes, Pilot
from flights.utils.excel_import_share import find_svodnaya_sheet_name, resolve_import_file, sheet_last_data_row


class Command(BaseCommand):
    help = 'Обновляет result_raw и result у вылетов из колонки O сводной Excel.'

    COL_TIME = 2
    COL_RESULT = 15
    COL_FLIGHT_DATE = 18
    COL_FLIGHT_NUMBER = 19
    COL_OPERATOR_CALLNAME = 22
    BATCH_SIZE = 500
    START_ROW = 5

    def add_arguments(self, parser):
        parser.add_argument('--file', dest='file_path', help='Путь к сводной Excel.')
        parser.add_argument('--dry-run', action='store_true')

    def handle(self, *args, **options):
        try:
            file_path = options.get('file_path') or str(resolve_import_file())
        except (FileNotFoundError, ValueError) as exc:
            raise CommandError(str(exc)) from exc

        pilots = {pilot.callname.lower(): pilot for pilot in Pilot.objects.all()}
        pending: dict[tuple, str] = {}

        try:
            if str(file_path).endswith('.xlsm'):
                wb = openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)
            else:
                wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception:
            wb = openpyxl.load_workbook(file_path, data_only=True)

        sheet_name = find_svodnaya_sheet_name(wb)
        if sheet_name is None:
            wb.close()
            raise CommandError("Лист 'СВОДНАЯ' не найден в файле импорта.")

        ws = wb[sheet_name]
        last_row = sheet_last_data_row(ws, start_row=self.START_ROW)
        self.stdout.write(f'Чтение строк {self.START_ROW}..{last_row} листа «{sheet_name}»')

        idx_time = self.COL_TIME - self.COL_TIME
        idx_result = self.COL_RESULT - self.COL_TIME
        idx_date = self.COL_FLIGHT_DATE - self.COL_TIME
        idx_number = self.COL_FLIGHT_NUMBER - self.COL_TIME
        idx_pilot = self.COL_OPERATOR_CALLNAME - self.COL_TIME

        for row in ws.iter_rows(
            min_row=self.START_ROW,
            max_row=last_row,
            min_col=self.COL_TIME,
            max_col=self.COL_OPERATOR_CALLNAME,
            values_only=True,
        ):
            result_raw = row[idx_result]
            result_text = str(result_raw).strip() if result_raw else ''
            if not result_text:
                continue

            pilot_raw = row[idx_pilot]
            pilot_callname = str(pilot_raw).strip().lower() if pilot_raw else ''
            pilot = pilots.get(pilot_callname)
            if not pilot:
                continue

            date_value = row[idx_date]
            if isinstance(date_value, datetime.datetime):
                flight_date = date_value.date()
            elif isinstance(date_value, datetime.date):
                flight_date = date_value
            else:
                continue

            flight_number_raw = row[idx_number]
            if flight_number_raw is None or str(flight_number_raw).strip() == '':
                unique_str = f'{flight_date.isoformat()}_{pilot.id}_{len(pending)}'
                hash_value = int(hashlib.md5(unique_str.encode()).hexdigest()[:8], 16)
                flight_number = abs(hash_value) % (10 ** 8)
            else:
                try:
                    flight_number = int(float(flight_number_raw))
                except (TypeError, ValueError):
                    continue

            time_value = row[idx_time]
            if isinstance(time_value, datetime.datetime):
                flight_time = time_value.time()
            elif isinstance(time_value, datetime.time):
                flight_time = time_value
            elif time_value:
                try:
                    flight_time = datetime.datetime.strptime(str(time_value).strip(), '%H:%M:%S').time()
                except ValueError:
                    try:
                        flight_time = datetime.datetime.strptime(str(time_value).strip(), '%H:%M').time()
                    except ValueError:
                        continue
            else:
                continue

            pending[(flight_number, pilot.id, flight_date, flight_time)] = result_text

        wb.close()
        self.stdout.write(f'Строк в Excel с результатом: {len(pending)}')

        flight_dates = {key[2] for key in pending}
        to_update: list[Flight] = []
        for flight in (
            Flight.objects.filter(flight_date__in=flight_dates)
            .only('id', 'number', 'pilot_id', 'flight_date', 'flight_time', 'result_raw', 'result')
            .iterator(chunk_size=2000)
        ):
            key = (flight.number, flight.pilot_id, flight.flight_date, flight.flight_time)
            result_text = pending.get(key)
            if not result_text:
                continue
            new_result = FlightResultTypes.from_excel_text(result_text)
            if (flight.result_raw or '') == result_text and flight.result == new_result:
                continue
            flight.result_raw = result_text
            flight.result = new_result
            to_update.append(flight)

        if options['dry_run']:
            self.stdout.write(f'Будет обновлено записей: {len(to_update)}')
            return

        updated = 0
        for start in range(0, len(to_update), self.BATCH_SIZE):
            batch = to_update[start:start + self.BATCH_SIZE]
            Flight.objects.bulk_update(batch, ['result_raw', 'result'], batch_size=self.BATCH_SIZE)
            updated += len(batch)

        self.stdout.write(self.style.SUCCESS(f'Обновлено result_raw у {updated} записей.'))
