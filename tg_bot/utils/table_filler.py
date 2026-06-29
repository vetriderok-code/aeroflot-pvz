from sys import exception

from openpyxl.workbook import Workbook
import logging
from db_handler.db_class import PostgresHandler
from openpyxl import load_workbook
import datetime


logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

db = PostgresHandler()

def add_row_without_saving(ws, data: list):
    if data:
        ws = ws
        try:
            x, y = data[3].replace('X', '').replace('Y', '').replace('=', '').strip().split(' ')
        except Exception as e:
            x, y = '-', '-'

        if data[8] == 'defeated':
            result = 'поражено'
        elif data[8] == 'destroyed':
            result = 'уничтожено'
        else:
            result = 'не поражено'

        if data[6]:
            exp_typ = data[6].replace(' ', '')
        else:
            exp_typ = ''

        if data[7]:
            exp_dev = data[7].replace(' ', '')
        else:
            exp_dev = ''

        row = [' ', data[0], 'патрулирование', '', '', '', data[4], data[5], 1, exp_typ, 1,
               exp_dev, 'уничтожение', '', '2 ОА', data[9], data[10], data[11].strftime('%H:%M'), ]

        free_row = 5
        for i in range(5, ws.max_row):
            if ws.cell(row=i, column=2).value is None:
                free_row = i
                break
        for i in range(1, len(row)):
            ws.cell(free_row, i + 1, row[i])

def make_doc(day: datetime.date):
    now_time = datetime.datetime.now().time()
    time = datetime.time(5, 00)
    yesterday = day - datetime.timedelta(days=1)
    if now_time < time :
        query = """
        SELECT flight.flight_time, flight.target, flight.corrective, flight.coordinates, flight.direction,
            flight.drone, flight.explosive_type, flight.explosive_device, flight.result, pilot.callname,
            flight.number, flight.created
        FROM pilot
        LEFT JOIN flight ON flight.pilot_id = pilot.id
        WHERE flight.created > %s and flight.created < %s
        ORDER BY flight.created;"""
        data = db.fetch(query, (datetime.datetime(yesterday.year, yesterday.month, yesterday.day, time.hour, time.minute),
                                datetime.datetime(day.year, day.month, day.day, time.hour, time.minute)))
    else:
        query = """
        SELECT flight.flight_time, flight.target, flight.corrective, flight.coordinates, flight.direction,
            flight.drone, flight.explosive_type, flight.explosive_device, flight.result, pilot.callname,
            flight.number, flight.created
        FROM pilot
        LEFT JOIN flight ON flight.pilot_id = pilot.id
        WHERE flight.created > %s and flight.created < %s
        ORDER BY flight.created;"""
        data = db.fetch(query, (datetime.datetime(day.year, day.month, day.day, time.hour, time.minute),
                                datetime.datetime(day.year, day.month, day.day + 1, time.hour, time.minute)))
    wb = load_workbook(filename="./templates/main.xlsx")
    ws = wb[wb.sheetnames[1]]
    if data:
        for row in data:
            add_row_without_saving(ws, row)

    wb.save(filename=f"./Отчет_{day.strftime('%Y_%m_%d')}.xlsx")

